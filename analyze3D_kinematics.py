import os
import numpy as np
import pandas as pd
import seaborn as sns
import matplotlib.pyplot as plt
from natsort import natsorted
import h5py
from sklearn.tree import DecisionTreeClassifier
from dateutil.parser import parse
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl import load_workbook
from datetime import datetime
from scipy.optimize import curve_fit
from sklearn.linear_model import LinearRegression
from sklearn.preprocessing import PolynomialFeatures
from statistics import median
from scipy.ndimage import gaussian_filter1d
from scipy.signal import find_peaks, peak_widths
from scipy.stats import gaussian_kde
from scipy.ndimage import gaussian_filter1d
import warnings
import re


warnings.filterwarnings(action="ignore", category=RuntimeWarning)

"""
These functions are responsible for preprocessing of angle data and 3D pose data
"""
def exponential_moving_average(data, alpha):
    if isinstance(data, pd.Series):
        data = data.tolist()
    ema = [data[0]]
    for i in range(1, len(data)):
        ema.append(alpha * data[i] + (1 - alpha) * ema[-1])
    return ema
def find_derivative(y):
    return np.gradient(y, 1)
def determine_smooth_amplify_curve(data, sig, amp):
    if amp:
        data = np.asarray(data)
        data = gaussian_filter1d(data, sigma=sig)
        return np.asarray(data)
    else:
        return np.asarray(data)
def TraceData_Preprocessing(angle_data, kinematic_data, start, end):
    DataToBeShown = dict()
    trial_angle_data = angle_data[start:end]
    trial_kinematic_data = kinematic_data[start:end]

    distance_pltf_RmCT = CalculateDistance_Platform_RmCT(trial_kinematic_data, "platform", "R-mCT")
    Platform_error = trial_kinematic_data["platform_error"]
    Platform_x = trial_kinematic_data["platform_x"]
    Platform_y = trial_kinematic_data["platform_y"]
    Platform_z = trial_kinematic_data["platform_z"]
    RightWing = trial_angle_data["RightWingAngle"]
    LeftWing = trial_angle_data["LeftWingAngle"]
    # CTAngle = trial_angle_data["CTAngle"]
    # CT_axis = trial_angle_data["CTAngle_axis"]
    # CT_cross = trial_angle_data["CTAngle_cross"]
    # FTAngle = trial_angle_data["FTAngle"]

    distance_pltf_RmCT = determine_smooth_amplify_curve(distance_pltf_RmCT, 5, True)
    Platform_x = determine_smooth_amplify_curve(Platform_x, 10, True)
    Platform_y = determine_smooth_amplify_curve(Platform_y, 10, True)
    Platform_z = determine_smooth_amplify_curve(Platform_z, 10, True)
    # CTAngle = determine_smooth_amplify_curve(CTAngle, True)
    # CT_axis = determine_smooth_amplify_curve(CT_axis, True)
    # CT_cross = determine_smooth_amplify_curve(CT_cross, True)
    # FTAngle = determine_smooth_amplify_curve(FTAngle, True)
    RightWing = determine_smooth_amplify_curve(RightWing, 10, True)
    LeftWing = determine_smooth_amplify_curve(LeftWing, 10, True)

    # platform_distance = determine_smooth_amplify_curve(trial_kinematic_data["platform_x"], True)
    DataToBeShown["platform"] = pd.Series(distance_pltf_RmCT)
    DataToBeShown["RightWingAngle"] = pd.Series(RightWing)
    DataToBeShown["LeftWingAngle"] = pd.Series(LeftWing)
    # DataToBeShown["CTAngle"] = pd.Series(CTAngle)
    # DataToBeShown["CT_axis"] = pd.Series(CT_axis)
    # DataToBeShown["CT_cross"] = pd.Series(CT_cross)
    # DataToBeShown["FTAngle"] = pd.Series(FTAngle)
    DataToBeShown["Platform_error"] = pd.Series(Platform_error)
    DataToBeShown["Platform_x"] = pd.Series(Platform_x)
    DataToBeShown["Platform_y"] = pd.Series(Platform_y)
    DataToBeShown["Platform_z"] = pd.Series(Platform_z)

    return DataToBeShown
def CalculateDistance_Platform_RmCT(threeD_data, keypoint1, keypoint2):
    platform_coxa_distance = []
    """

    for p_x, p_y, p_z, C_x, C_y, C_z in zip(threeD_data[f"{keypoint1}_x"], threeD_data[f"{keypoint1}_y"],
                                            threeD_data[f"{keypoint1}_z"], threeD_data[f"{keypoint2}_x"],
                                            threeD_data[f"{keypoint2}_y"], threeD_data[f"{keypoint2}_z"]):
        platform_coxa_distance.append(np.sqrt((p_x - C_x) ** 2 + (p_y - C_y) ** 2 + (p_z - C_z) ** 2))
    # sns.lineplot(x=range(len(threeD_data)), y=platform_coxa_distance)
    return platform_coxa_distance
    """
    C_x = np.average(threeD_data[f"{keypoint2}_x"][:100])
    C_y = np.average(threeD_data[f"{keypoint2}_y"][:100])
    C_z = np.average(threeD_data[f"{keypoint2}_z"][:100])
    for p_x, p_y, p_z in zip(threeD_data[f"{keypoint1}_x"], threeD_data[f"{keypoint1}_y"],
                                            threeD_data[f"{keypoint1}_z"]):
        platform_coxa_distance.append(np.sqrt((p_x - C_x) ** 2 + (p_y - C_y) ** 2 + (p_z - C_z) ** 2))
    # sns.lineplot(x=range(len(threeD_data)), y=platform_coxa_distance)
    return platform_coxa_distance
def normalize_list(data, method="min-max"):
    if method == "min-max":
        min_val = min(data)
        max_val = max(data)
        if max_val == min_val:
            raise ValueError("Cannot perform Min-Max normalization when all values are the same.")
        return [(x - min_val) / (max_val - min_val) for x in data]
"""
These functions are responsible for detecting various characteristic of the angle and 3D data
"""
def detect_brief_landing(angle_data, bl_range, wing_ang_fold_th, wing_ang_drop_percent, window_size, st_base_th, consecutive_fold_th):

    # Calculate the start and end baseline of wing angle
    start_baseline = np.mean(angle_data[:bl_range])
    end_baseline = np.mean(angle_data[-bl_range:])

    chunk_size = 10
    consecutive_chunk = 0

    if start_baseline > st_base_th:

        for i in range(0, len(angle_data), window_size):
            curr_v = np.mean(angle_data[i:i + window_size])
            if curr_v / start_baseline < (1 - wing_ang_drop_percent) and curr_v < wing_ang_fold_th:
                consecutive_chunk += 1
            else:
                consecutive_chunk = 0
            if consecutive_chunk >= consecutive_fold_th:
                stop_idx = i - (consecutive_fold_th - 1) * window_size
                return True, stop_idx
    return False, 0
def detect_significant_drop(angle_data, bl_range, wing_ang_fold_th, wing_ang_drop_percent, window_size):
    start_baseline = np.mean(angle_data[:bl_range])
    end_baseline = np.mean(angle_data[-bl_range:])
    repeat = 0
    cons_th = 4
    if (1 - (end_baseline / start_baseline)) > 0.05:
        # print(start_baseline)
        for i in range(0, len(angle_data), window_size):
            current_chunk_v = np.mean(angle_data[i:i + window_size])
            if (abs(current_chunk_v - start_baseline) / (start_baseline - end_baseline) > wing_ang_drop_percent) and current_chunk_v < wing_ang_fold_th:
                # return i

                repeat += 1
            else:
                repeat = 0
            if repeat >= cons_th:
                return i - ((cons_th - 1) * window_size)


    return 0
def detect_Not_Flying(angle_data, bl_range, wing_ang_fold_th=0, wing_ang_drop_percent=0, window_size=1, baseline_diff_percent=0):
    start_base_line = np.mean(angle_data[:bl_range])
    end_base_line = np.mean(angle_data[-bl_range:])
    if start_base_line < 2:
        return True
    else:
        return False
def detect_moment_of_landing(trace_data, wing_distance_threshold, wind_size, duration_threshold):
    wing_fold_duration = 0
    cons_th = 4

    for i in range(0, len(trace_data), wind_size):
        if np.average(trace_data[i:i+wind_size]) < wing_distance_threshold:
            wing_fold_duration += 1
        else:
            wing_fold_duration = 0
        if wing_fold_duration > duration_threshold:
            return i - ((cons_th - 1) * wind_size), True
    return 0, False
def detectFlying(angle_data, bl_range, wing_ang_fold_th, window_size):
    chunk_size = 20
    start_base_line = np.mean(angle_data[:bl_range])
    end_base_line = np.mean(angle_data[-bl_range:])
    if start_base_line > wing_ang_fold_th and end_base_line > wing_ang_fold_th:
        for i in range(0, len(angle_data), window_size):
            current_v = np.mean(angle_data[i:i + window_size])
            if current_v < wing_ang_fold_th:
                return False
    return True
def detect_moment_of_landing_WingAngle(angle_data, wing_angle_threshold, wind_size, duration_threshold):
    wing_fold_duration = 0
    cons_th = 3

    for i in range(0, len(angle_data), wind_size):
        if np.average(angle_data[i:i+wind_size]) < wing_angle_threshold:
            wing_fold_duration += 1
        else:
            wing_fold_duration = 0
        if wing_fold_duration > duration_threshold:
            return i - ((cons_th - 1) * wind_size), True
    return 0, False
"""
These functions are responsible for testing and plotting the result from detection functions.
"""
def ReadAndFilterData(GroupName, Flies_to_pick, Landing_Data_path):
    global Trial_num
    Landing_Data = pd.read_excel(Landing_Data_path)
    Landing_Data = Landing_Data.iloc[Flies_to_pick[0] - 1:Flies_to_pick[1]]
    Valid_data_index = []
    for index, row in Landing_Data.iterrows():
        str_nan_count = 0
        for data in row.values:
            if isinstance(data, str) or pd.isna(data):
                str_nan_count += 1
        if str_nan_count < (len(row.values) / 2):
            Valid_data_index.append(index)

    Landing_Data = Landing_Data[Landing_Data.index.isin(Valid_data_index)]
    GroupNameCol = [GroupName] * len(Valid_data_index)
    Landing_Data["Group_Name"] = GroupNameCol
    return Landing_Data
def CalculateLPAndmLLAcrossFlies(GroupName, Landing_Data):
    global FPS
    LP_mLL_Data = dict()
    LP_mLL_Data["LandingProb"] = []
    LP_mLL_Data["MLandingLatency"] = []
    LP_mLL_Data["Fly#"] = []
    # LP_mLL_Data["Time"] = Landing_Data["Time"].tolist()
    # LP_mLL_Data["Date"] = Landing_Data["Date"].tolist()
    Trials = ["Trial_" + str(i + 1) for i in range(Trial_num)]
    Landing_Data = Landing_Data[Trials]
    for index, row in Landing_Data.iterrows():
        Nan_data = [n for n in row if pd.isna(n) or isinstance(n, str)]
        Flying = [f for f in row if f == -1]
        Landing_latency = [l / 250 for l in row if not isinstance(l, str) and l > -1]
        if len(Nan_data) + len(Flying) + len(Landing_latency) != Trial_num:
            print(f"Error while filtering data")
            print(f"Index: {index}")
            print(f"# of Nan: {len(Nan_data)}\t{Nan_data}")
            print(f"# of Flying: {len(Flying)}\t{Flying}")
            print(f"# of Landing: {len(Landing_latency)}\t{Landing_latency}")
        LP_mLL_Data["Fly#"].append(index + 1)
        LP_mLL_Data["LandingProb"].append(len(Landing_latency) / (len(Flying) + len(Landing_latency)))
        LP_mLL_Data["MLandingLatency"].append(np.mean(Landing_latency))
    LP_mLL_Data["Group_Name"] = [GroupName] * Landing_Data.shape[0]
    LP_mLL_Data = pd.DataFrame(LP_mLL_Data)
    return LP_mLL_Data
def LLCluster(Original_Data, Filter_Data):
    global FPS
    global markers
    Trials = ["Trial_" + str(i + 1) for i in range(Trial_num)]
    fig = plt.figure(1, figsize=(16, 8))
    sns.axes_style("ticks")
    for i, group_data in enumerate(Original_Data, start=1):
        Landing_latency = []
        n = 0
        print(group_data)
        for index, fly in group_data[Trials].iterrows():
            Landing_latency.extend([l / 250 for l in fly if not isinstance(l, str) and l > -1])

            # sns.stripplot(x=np.full(len(Landing_latency), i + 2), y=Landing_latency, jitter=0.2, marker='o', size=8, alpha=0.4, color=colors[n])
        sns.violinplot(x=np.full(len(Landing_latency), i), y=Landing_latency, color=Color_blind_palette[i - 1],
                       fill=False)

    for i, group_data in enumerate(Filter_Data, start=1):
        # ax = sns.stripplot(x="Group_Name", y="MLandingLatency", data=group_data, jitter=0.2, marker=markers[i - 1], size=20, alpha=0.75, color=Color_blind_palette[i-1])
        ax = sns.violinplot(x="Group_Name", y="MLandingLatency", data=group_data, color=Color_blind_palette[i - 1])
        ax.spines['left'].set_linewidth(3)
        ax.spines['bottom'].set_linewidth(3)
    plt.ylabel("Landing latency (s)", fontsize=25)

    plt.ylim(0, 3.2)
    plt.yticks([0, 1.5, 3])
    plt.xticks(ticks=[0, 1, 2, 3],
               labels=['Trial LL M', 'Trial LL C', 'Mean LL M', 'Mean LL C'])
    plt.tick_params(axis="y", labelsize=25)
    plt.tick_params(axis="x", labelsize=20)
    plt.tick_params(width=3, length=10)
    sns.despine()
    # fig.suptitle("Landing latency (T2-TTa vs T2-CTF)")
    # plt.rcParams["axes.linewidth"] = 2.5
    plt.tight_layout()

    plt.savefig("LLAcrossFlies.pdf", dpi=300, bbox_inches='tight')
    plt.show()
def LPAcrossFlies(Data_to_plot):
    global Color_blind_palette
    global markers

    combined_df = pd.concat(Data_to_plot)
    plt.figure(figsize=(16, 8))

    for i, d in enumerate(Data_to_plot):
        print(i)
        # g = sns.stripplot(x="Group_Name", y="LandingProb", hue="Group_Name", data=d, marker=markers[i], jitter=0.3, alpha=0.75, size=20, palette=[Color_blind_palette[i]])
        g = sns.violinplot(x="Group_Name", y="LandingProb", hue="Group_Name", data=d, palette=[Color_blind_palette[i]])
        g.spines['left'].set_linewidth(3)
        g.spines['bottom'].set_linewidth(3)
        g.set_ylim(-0.1, 1.3)

    group_stat = combined_df.groupby('Group_Name')['LandingProb'].agg(['mean', 'std', 'count']).reset_index()
    group_stat['ci'] = 1.96 * group_stat['std'] / np.sqrt(group_stat['count'])

    sns.pointplot(x='Group_Name', y='mean', data=group_stat, color='black', linestyles=" ", markers="s", errorbar=None)
    plt.errorbar(x=group_stat['Group_Name'], y=group_stat['mean'], yerr=group_stat['ci'], fmt="none", color='black', capsize=5)

    # plt.legend(handles=legend_elements, loc='upper right', fontsize=20)
    plt.ylabel("Landing Probability", fontsize=25)
    plt.tick_params(axis="y", labelsize=25)
    plt.tick_params(axis="x", labelsize=25)
    plt.tick_params(width=3, length=10)
    plt.yticks([0, 0.5 ,1])

    sns.despine()
    plt.savefig("LPAcrossFliesStarvedExp", dpi=300, bbox_inches='tight')
    plt.show()
def CalculatePlatform_distance_average(collected_data):
    np_matrix = np.array(collected_data)
    transposed_matrix = np_matrix.T
    average_trace = []
    for frame_data in transposed_matrix:
        average_trace.append(np.average(frame_data))
    # average_trace = determine_smooth_amplify_curve(average_trace, 5, True)
    return average_trace
def detectMOC(data, threshold=0.09538):
    wind_size = 5
    baseline = np.average(data[400:])
    for i in range(0, len(data), wind_size):
        if np.average(data[i:i + wind_size]) <= threshold:
            return i
    return 0
def MakePlot(DataPaths, fly_num):
    i = 0
    j = 0
    t = 0
    start = 0
    end = 1250
    fig, axs = plt.subplots(nrows=5, ncols=4, figsize=(20, 16))
    for dp in DataPaths:
        Data = pd.read_csv(dp)
        if i % 5 == 0:
            j += 1
            i = 1
        else:
            i += 1
        # Plot data
        ax2 = axs[i - 1, j - 1].twinx()
        ax2.set_ylim(0, 180)

        trial_type, MOL = CategorizeTrial(Data, fly_num, t + 1)

        WingAbTipAngle = Calculate_joint_angle(Data, [["R-wing", "R-wing-hinge", "abdomen-tip"], ["L-wing", "L-wing-hinge", "abdomen-tip"]])
        WingTipDistance = Calculate_segment_length(Data, [["L-wing", "R-wing"]])
        R_mol, RL = detect_moment_of_landing_WingAngle(WingAbTipAngle["R-wing-hinge"], 50, 4, 2)
        L_mol, LL = detect_moment_of_landing_WingAngle(WingAbTipAngle["L-wing-hinge"], 50, 4, 2)
        # mol, L = detect_moment_of_landing(WingTipDistance["L-wing_R-wing"], 2, 4, 3)
        seconds = [f for f in range(len(WingTipDistance["L-wing_R-wing"][start:end]))]
        # sns.lineplot(x=seconds, y=WingTipDistance["L-wing_R-wing"][start:end], ax=axs[i - 1, j - 1], color="black")
        # sns.lineplot(x=seconds, y=trial_angle["LeftWingAngle"][clipstart:clipend], ax=axs[i - 1, j - 1], color="blue")
        # sns.lineplot(x=seconds, y=wing_distance, ax=ax2, color="green")

        if trial_type == 0:
            sns.lineplot(x=seconds, y=WingAbTipAngle["R-wing-hinge"][start:end], ax=ax2, color="blue")
            sns.lineplot(x=seconds, y=WingAbTipAngle["L-wing-hinge"][start:end], ax=ax2, color="navy")
        elif trial_type == 1:
            MOC, contact_point = DetermineTiTaMOC(Data, fly_num, t + 1)
            if contact_point == "NoContact" or contact_point == "L":
                sns.lineplot(x=seconds, y=WingAbTipAngle["R-wing-hinge"][start:end], ax=ax2, color="red")
                sns.lineplot(x=seconds, y=WingAbTipAngle["L-wing-hinge"][start:end], ax=ax2, color="maroon")
            else:
                sns.lineplot(x=seconds, y=WingAbTipAngle["R-wing-hinge"][start:end], ax=ax2, color="yellow")
                sns.lineplot(x=seconds, y=WingAbTipAngle["L-wing-hinge"][start:end], ax=ax2, color="gold")
                axs[i - 1, j - 1].axvline(R_mol - start, color="black")
                axs[i - 1, j - 1].axvline(L_mol - start, color="grey")
                axs[i - 1, j - 1].axvline(MOC - start, color="red")
        elif trial_type == -1:
            MOC, contact_point = DetermineTiTaMOC(Data, fly_num, t + 1)
            if contact_point == "NoContact" or contact_point == "L":
                sns.lineplot(x=seconds, y=WingAbTipAngle["R-wing-hinge"][start:end], ax=ax2, color="red")
                sns.lineplot(x=seconds, y=WingAbTipAngle["L-wing-hinge"][start:end], ax=ax2, color="maroon")
            else:
                sns.lineplot(x=seconds, y=WingAbTipAngle["R-wing-hinge"][start:end], ax=ax2, color="darkorange")
                sns.lineplot(x=seconds, y=WingAbTipAngle["L-wing-hinge"][start:end], ax=ax2, color="orange")

        axs[i - 1, j - 1].set_xlabel("Frames")
        # axs[i - 1, j - 1].set_ylabel("Platform distance (mm)")
        axs[i - 1, j - 1].set_yticks([0, 2, 4, 6, 8])
        axs[i - 1, j - 1].set_ylim(-0.1, 8.1)
        axs[i - 1, j - 1].set_ylabel("Distance (mm)")
        # axs[i - 1, j - 1].set_yticks([0, 50, 100, 150, 200])
        axs[i - 1, j - 1].set_title(f"Trial {t + 1}")
        t += 1
    print(f"Plotting {fly_num}'s data")
    plt.suptitle(str(fly_num), fontsize=20)
    plt.tight_layout()  # Adjust subplot layout
    plt.savefig("KirExp_Ti-Ta" + str(fly_num) + ".png")  # Save the figure
    plt.close()
def PlotWingAngleDistribution(AngleData, kinematic_pose_data):
    global Analysis_start
    global Analysis_end
    global Customized_wing_threshold
    i = 0
    j = 0
    t = 0
    fig, axs = plt.subplots(nrows=5, ncols=6, figsize=(20, 16))
    for key in AngleData.keys():
        if i % 5 == 0:
            j += 1
            i = 1
        else:
            i += 1
        R_wing = []
        L_wing = []
        for trial_angle, trial_kinematic in zip(AngleData[key], kinematic_pose_data[key]):
            DataToBeShown = TraceData_Preprocessing(trial_angle, trial_kinematic, Analysis_start, Analysis_end)
            #moment_of_contact_manual = platform_moc.iloc[int(fly) - 1][f"Trial_{t + 1}"]
            #moment_of_landing = mol_data.iloc[int(fly) - 1][f"Trial_{t + 1}"]

            nan_count = DataToBeShown["RightWingAngle"].isna().sum()

            percentage_nan = (nan_count / len(AngleData)) * 100
            moment_of_contact = 0
            stop_dropping = 0
            # Determine validity of the data
            # print(f"Fly: {fly}---Trial{trial}")
            if percentage_nan < 50:
                R_wing.extend(DataToBeShown["RightWingAngle"])
                L_wing.extend(DataToBeShown["LeftWingAngle"])
                # sns.lineplot(x=[i / 250 for i in range(len(DataToBeShown["RightWingAngle"]))], y=DataToBeShown["RightWingAngle"], ax=axs[i - 1, j - 1], color="orange")
                # sns.lineplot(x=[i / 250 for i in range(len(DataToBeShown["RightWingAngle"]))], y=DataToBeShown["LeftWingAngle"], ax=axs[i - 1, j - 1], color="blue")

        if not len(R_wing) == 0:
            sns.histplot(R_wing, binwidth=2, ax=axs[i - 1, j - 1], color="orange", kde=True)
            sns.histplot(L_wing, binwidth=2, ax=axs[i - 1, j - 1], color="blue", kde=True)
            bin_width = 2

            # Calculate the number of bins
            bin_range = np.arange(min(R_wing), max(R_wing) + bin_width, bin_width)

            # Calculate histogram
            hist, bin_edges = np.histogram(R_wing, bins=bin_range)
            max_f = 0
            # Print bin edges and frequencies

            # axs[i - 1, j - 1].axvline(Customized_wing_threshold[int(key) - 1], color="black")

        axs[i - 1, j - 1].set_title(f"Fly {key}")
        # axs[i - 1, j - 1].set_ylabel("Wing Angle")
        # axs[i - 1, j - 1].set_xlabel("Time (s)")
        axs[i - 1, j - 1].set_xlabel("Wing Angle")
        print(f"Plotting {key} data")
        # axs[i - 1, j - 1].set_xticks([50, 100, 150, 200])
        # axs[i - 1, j - 1].set_ylim(0, 350)
    # plt.suptitle(f"Fly {key}")
    plt.tight_layout()
    # print(f"Plotting fly {key} data")
    plt.savefig("Trace_Superimposed.png")  # Save the figure
    plt.close()
# def FTAngleChangeThroughTrials(DataPaths):

"""
These functions are responsible for extracting and writing the data from and to CSV file.
"""
# Extract the data path of the csv files and organize them in the order by fly number
def Analyze_3D_pose_data(FlyDataPath, Fly_num):
    Landing_Data = []
    LandingLatency_Data = []
    for f, k in enumerate(list(FlyDataPath.keys())[:Fly_num]):
        Fly_Landing_Data = []
        Fly_LandingLatency_Data = []
        for t, trial_data_path in enumerate(FlyDataPath[k]):
            pose_data = pd.read_csv(trial_data_path)
            trial_type, MOL = CategorizeTrial(pose_data, f + 1, t + 1)
            if trial_type == 0:
                Fly_LandingLatency_Data.append("NF")
            elif trial_type == 1:
                MOC, contact_point = DetermineTiTaMOC(pose_data, f + 1, t + 1)
                if contact_point == "NoContact" or contact_point == "L":
                    Fly_LandingLatency_Data.append(np.nan)
                else:
                    Fly_LandingLatency_Data.append(MOL - MOC)
                    # Fly_LandingLatency_Data.append(MOL)
            elif trial_type == -1:
                MOC, contact_point = DetermineTiTaMOC(pose_data, f + 1, t + 1)
                if contact_point == "NoContact" or contact_point == "L":
                    Fly_LandingLatency_Data.append(np.nan)
                else:
                    Fly_LandingLatency_Data.append(-1)
        LandingLatency_Data.append(Fly_LandingLatency_Data)
    return LandingLatency_Data
def Write_to_csv(LandingProbData, LandingLatencyData, output_path, compare_path):
    Manual_data = pd.read_excel(compare_path)
    T = ["Trial_" + str(i) for i in range(1, 21)]
    Manual_data = Manual_data[T]
    # Create an Excel writer using openpyxl
    excel_writer = pd.ExcelWriter(output_path, engine='openpyxl')
    LandingProbData.to_excel(excel_writer, index=False, sheet_name='Sheet1')
    # Access the workbook and sheet
    workbook = excel_writer.book
    sheet = workbook['Sheet1']

    # Define cell colors
    yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
    orange_fill = PatternFill(start_color='FFC000', end_color='FFC000', fill_type='solid')
    red_fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
    blue_fill = PatternFill(start_color='9BC2E6', end_color='9BC2E6', fill_type='solid')
    gray_fill = PatternFill(start_color='696969', end_color='696969', fill_type='solid')

    # Apply cell colors based on conditions
    for row_idx in range(2, sheet.max_row + 1):  # Start from row 2 (excluding header)
        for col_idx in range(2, sheet.max_column + 1):  # Start from column 2 (excluding Fly# column)
            cell_value = sheet.cell(row=row_idx, column=col_idx).value
            cell = sheet.cell(row=row_idx, column=col_idx)
            if cell_value == 1:
                if isinstance(Manual_data[T].iloc[row_idx - 2][col_idx - 2], str):
                    cell.fill = gray_fill
                else:
                    if Manual_data[T].iloc[row_idx - 2][col_idx - 2] > 1:
                        cell.fill = yellow_fill
                    else:
                        cell.fill = gray_fill
                cell.value = LandingLatencyData[row_idx - 2][col_idx - 2]
            elif cell_value == -1:
                if Manual_data[T].iloc[row_idx - 2][col_idx - 2] == -1:
                    cell.fill = orange_fill
                else:
                    cell.fill = gray_fill
            elif cell_value == 0:
                if isinstance(Manual_data[T].iloc[row_idx - 2][col_idx - 2], str):
                    if Manual_data[T].iloc[row_idx - 2][col_idx - 2] == "NotFlying":
                        cell.fill = blue_fill
                    else:
                        cell.fill = gray_fill
                else:
                    cell.fill = gray_fill
                cell.value = "NotFlying"
            elif isinstance(type(cell_value), type(str)):
                # print("N/A")
                # print(cell_value)
                if pd.isna(Manual_data[T].iloc[row_idx - 2][col_idx - 2]):
                    cell.fill = red_fill
                else:
                    cell.fill = gray_fill
                cell.value = 'N/A'
    # Save the Excel file
    excel_writer.save()
    print("Excel file has been created with modifications.")
def OutptuPrediction(Predicted_Data, outputpath):
    df_transformed = pd.DataFrame(Predicted_Data, columns=[f"Trial_{i + 1}" for i in range(20)])
    df_transformed.insert(0, "Fly#", range(1, len(df_transformed) + 1))

    # Replace NaN values with "N/A" for display
    df_transformed.fillna("N/A", inplace=True)

    # Save to an Excel file
    output_file = f"{outputpath}.xlsx"
    df_transformed.to_excel(output_file, index=False)

    # Load the workbook for formatting
    wb = load_workbook(output_file)
    ws = wb.active

    # Define color fills
    color_fills = {
        "N/A": PatternFill(start_color="FF6666", end_color="FF6666", fill_type="solid"),  # Red for N/A
        "NF": PatternFill(start_color="6DB6FF", end_color="6DB6FF", fill_type="solid"),  # Blue for NF
        "-1": PatternFill(start_color="FF9933", end_color="FF9933", fill_type="solid"),  # Orange for -1
        "number": PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")  # Yellow for other numbers
    }

    # Apply color coding (skip header row)
    for row in ws.iter_rows(min_row=2, min_col=2):  # Start from column 2 to skip "Fly#"
        for cell in row:
            if isinstance(cell.value, str) and cell.value == "N/A":
                cell.fill = color_fills["N/A"]
            elif isinstance(cell.value, str) and cell.value == "NF":
                cell.fill = color_fills["NF"]
            elif isinstance(cell.value, (int, float)) and cell.value == -1:
                cell.fill = color_fills["-1"]
            elif isinstance(cell.value, (int, float)):  # Any other number
                cell.fill = color_fills["number"]

    # Save the formatted workbook
    wb.save(output_file)

    print(f"File saved as {output_file}")
def Read_Manual_Data(file_path):
    Trial_num = 20
    Trial_num_list = ["Trial_" + str(i + 1) for i in range(Trial_num)]
    Manual_data = pd.read_excel(file_path)[Trial_num_list]
    return Manual_data
def group_files_by_fly(file_names):
    # Regular expression to extract the relevant parts of the file name
    pattern = re.compile(r"(\d{4}-\d{2}-\d{2})-\d{2}-\d{2}-\d{2}\.\d+.*?_Fly_(\d+)_Trial_\d+_")
    # Dictionary to map unique combinations to "Fly_N"
    unique_combinations = {}
    grouped_files = {}

    current_fly_number = 1  # Counter for Fly_N keys

    for file in file_names:
        match = pattern.search(file)
        if match:
            date = match.group(1)
            fly_number = match.group(2)
            unique_id = (date, fly_number)  # Unique identifier based on date and fly number

            # Assign a new Fly_N key if the combination is not seen before
            if unique_id not in unique_combinations:
                unique_combinations[unique_id] = f"Fly_{current_fly_number}"
                current_fly_number += 1

            # Get the Fly_N key
            fly_key = unique_combinations[unique_id]

            # Add the file to the corresponding Fly_N group
            if fly_key not in grouped_files:
                grouped_files[fly_key] = []
            grouped_files[fly_key].append(file)

    return grouped_files
def Calculate_distance_between_points(x, y, z, x1, y1, z1):
    return np.sqrt((x - x1) ** 2 + (y - y1) ** 2 + (z - z1) ** 2)
def Calculate_segment_length(threeD_data, skeletons):
    collected_seg_length_data = dict()
    for seg in skeletons:
        if f"{seg[0]}_{seg[1]}" not in collected_seg_length_data.keys():
            collected_seg_length_data[f"{seg[0]}_{seg[1]}"] = []
        for f in range(len(threeD_data)):
            collected_seg_length_data[f"{seg[0]}_{seg[1]}"].append(Calculate_distance_between_points(
                threeD_data[f"{seg[0]}_x"][f], threeD_data[f"{seg[0]}_y"][f], threeD_data[f"{seg[0]}_z"][f],
                threeD_data[f"{seg[1]}_x"][f], threeD_data[f"{seg[1]}_y"][f], threeD_data[f"{seg[1]}_z"][f]))
    collected_seg_length_data = pd.DataFrame(collected_seg_length_data)
    return collected_seg_length_data
def calculate_angle(x1, y1, z1, x2, y2, z2, x3, y3, z3):
    # Define points as numpy arrays
    pt1 = np.array([x1, y1, z1])
    pt2 = np.array([x2, y2, z2])
    pt3 = np.array([x3, y3, z3])

    # Define vectors
    vecA = pt1 - pt2
    vecB = pt3 - pt2

    # Calculate dot product and magnitudes
    dot_product = np.dot(vecA, vecB)
    magnitude_A = np.linalg.norm(vecA)
    magnitude_B = np.linalg.norm(vecB)

    # Calculate angle in radians
    angle_rad = np.arccos(dot_product / (magnitude_A * magnitude_B))

    # Convert to degrees if needed
    angle_deg = np.degrees(angle_rad)

    return angle_deg
def Calculate_joint_angle(threeD_data, angles):
    collected_angle_data = dict()
    for ag in angles:
        if f"{ag[1]}" not in collected_angle_data.keys():
            collected_angle_data[f"{ag[1]}"] = []
        for f in range(len(threeD_data)):
            angle = calculate_angle(
                threeD_data[f"{ag[0]}_x"][f], threeD_data[f"{ag[0]}_y"][f], threeD_data[f"{ag[0]}_z"][f],
                threeD_data[f"{ag[1]}_x"][f], threeD_data[f"{ag[1]}_y"][f], threeD_data[f"{ag[1]}_z"][f],
                threeD_data[f"{ag[2]}_x"][f], threeD_data[f"{ag[2]}_y"][f], threeD_data[f"{ag[2]}_z"][f])
            collected_angle_data[f"{ag[1]}"].append(angle)
    return collected_angle_data
def TransposeData(df):
    df = pd.DataFrame(df)
    df = df.T
    return df
def Calculate_derivative(data):
    if len(data) < 2:
        raise ValueError("The input list must contain at least two elements to calculate a derivative.")

    # Compute the differences between consecutive elements
    data = [data[i + 1] - data[i] for i in range(len(data) - 1)]
    data = [abs(x) for x in data]
    return data
def line_plane_intersection(p1, p2, normal, d):
    # Line represented as p(t) = p1 + t * (p2 - p1)
    direction = np.array(p2) - np.array(p1)
    denom = np.dot(normal, direction)

    if abs(denom) < 1e-6:  # Parallel case
        return None

    t = -(np.dot(normal, p1) + d) / denom

    if 0 <= t <= 1:  # Ensure intersection is within the segment
        return p1 + t * direction, t
    return None, None
def is_inside_circle(intersection, center, radius):
    return np.linalg.norm(intersection - center) <= radius
def check_leg_platform_intersection(leg_p1, leg_p2, center_traces, center_point, platform_offset):
    # Compute platform radius
    radius = 0.33
    # Compute the plane equation
    centroid, platform_direction = best_fit_line_3d(center_traces)

    platform_plane_origin = center_point + platform_offset * platform_direction  # Shift the plane up
    d = -np.dot(platform_direction, platform_plane_origin)

    # Find intersection point
    intersection, intersect_proportion = line_plane_intersection(np.array(leg_p1), np.array(leg_p2), platform_direction, d)

    if intersection is not None and is_inside_circle(intersection, np.array(platform_plane_origin), radius):
        return True, intersect_proportion
    return False, None
def ReadCoordsAll(threeD_data, fnum):
    points = ["L-wing", "L-wing-hinge", "R-wing", "R-wing-hinge", "abdomen-tip", "platform-tip", "L-platform-tip",
              "R-platform-tip", "platform-axis", "R-fBC", "R-fCT", "R-fFT", "R-fTT", "R-fLT", "R-mBC", "R-mCT", "R-mFT",
              "R-mTT", "R-mLT", "R-hBC", "R-hCT", "R-hFT", "R-hTT", "R-hLT", "L-fBC", "L-fCT", "L-fFT", "L-fTT", "L-fLT",
              "L-mBC", "L-mCT", "L-mFT", "L-mTT", "L-mLT", "L-hBC", "L-hCT", "L-hFT", "L-hTT", "L-hLT"]
    coords = dict()
    for p in points:
        coords[p] = np.asarray([threeD_data[f"{p}_x"][fnum], threeD_data[f"{p}_y"][fnum], threeD_data[f"{p}_z"][fnum]])
    return coords
def plot_motion_vector_with_plane(platform_ctr_pts_traces, current_center, coords, platform_offset):
    keypoint_pairs = [
        ["L-wing", "L-wing-hinge"], ["R-wing", "R-wing-hinge"], ["abdomen-tip"],
        ["platform-tip"], ["L-platform-tip"], ["R-platform-tip"], ["platform-axis"],
        ["R-fBC", "R-fCT", "R-fFT", "R-fTT", "R-fLT"], ["R-mBC", "R-mCT", "R-mFT", "R-mTT", "R-mLT"],
        ["R-hBC", "R-hCT", "R-hFT", "R-hTT", "R-hLT"], ["L-fBC", "L-fCT", "L-fFT", "L-fTT", "L-fLT"],
        ["L-mBC", "L-mCT", "L-mFT", "L-mTT", "L-mLT"], ["L-hBC", "L-hCT", "L-hFT", "L-hTT", "L-hLT"]
    ]

    platform_ctr_pts_traces = np.array(platform_ctr_pts_traces)
    centroid, direction = best_fit_line_3d(platform_ctr_pts_traces)
    # Generate points along the best-fit line
    t_vals = np.linspace(-10, 10, 100)
    line_points = centroid + np.outer(t_vals, direction)

    # Create color gradient based on the order of the points
    num_points = len(platform_ctr_pts_traces)
    colors = plt.cm.Greys(np.linspace(0, 1, num_points))

    # Generate a normal plane to the direction vector at the centroid
    normal_vector = direction
    # Create two perpendicular vectors using cross products
    if np.allclose(normal_vector, [1, 0, 0]):  # Handle edge case if aligned with x-axis
        perp_vector1 = np.cross(normal_vector, [0, 1, 0])
    else:
        perp_vector1 = np.cross(normal_vector, [1, 0, 0])

    perp_vector1 /= np.linalg.norm(perp_vector1)  # Normalize

    perp_vector2 = np.cross(normal_vector, perp_vector1)
    perp_vector2 /= np.linalg.norm(perp_vector2)  # Normalize

    radius = 0.32 # Radius of the circular plane

    # Generate a circular grid
    u_vals = np.linspace(-radius, radius, 50)
    v_vals = np.linspace(-radius, radius, 50)
    U, V = np.meshgrid(u_vals, v_vals)

    # Convert to polar coordinates to filter for a circular region
    mask = U ** 2 + V ** 2 <= radius ** 2  # Boolean mask for circle

    # Apply mask to keep only circular region
    U = U[mask]
    V = V[mask]

    platform_plane_origin = current_center + platform_offset * direction  # Shift the plane up

    plane_points = platform_plane_origin + U[..., None] * perp_vector1 + V[..., None] * perp_vector2

    # 3D Plot
    fig = plt.figure(figsize=(8, 6))
    ax = fig.add_subplot(111, projection='3d')

    # Loop through the connections and plot lines
    Colors = [
        "#FF0000", "#008000", "#0000FF", "#FFFF00", "#00FFFF", "#FF00FF", "#000000",
         "#808080", "#A9A9A9", "#D3D3D3", "#FFA500", "#800080", "#A52A2A",
        "#FFC0CB", "#ADD8E6", "#00FF00", "#4B0082", "#EE82EE", "#FFD700", "#C0C0C0",
        "#D2B48C", "#FF7F50", "#006400", "#00008B", "#8B0000", "#FF8C00", "#8B008B",
        "#708090", "#FF6347", "#008080"
    ]
    for g, group in enumerate(keypoint_pairs):
        for i in range(len(group) - 1):  # Connect points in the group
            p1 = coords[group[i]]
            p2 = coords[group[i + 1]]

            # Plot a line between p1 and p2
            ax.plot([p1[0], p2[0]], [p1[1], p2[1]], [p1[2], p2[2]], marker='o', color=Colors[g])

    ax.quiver(*platform_ctr_pts_traces[-1], *perp_vector1, color='r', arrow_length_ratio=0.1)
    ax.quiver(*platform_ctr_pts_traces[-1], *perp_vector2, color='r', arrow_length_ratio=0.1)
    # Plot the trajectory with color gradient
    for i in range(num_points - 1):
        ax.plot(platform_ctr_pts_traces[i:i + 2, 0], platform_ctr_pts_traces[i:i + 2, 1], platform_ctr_pts_traces[i:i + 2, 2], color=colors[i], linewidth=6)

    # Plot the best-fit line
    ax.plot(line_points[:, 0], line_points[:, 1], line_points[:, 2], 'r--', label="Best-Fit Line")

    # Highlight the centroid
    # ax.scatter(centroid[0], centroid[1], centroid[2], color='black', s=100, label="Centroid", edgecolors='k')

    # Plot the normal plane
    ax.plot_trisurf(plane_points[:, 0], plane_points[:, 1], plane_points[:, 2], color='cyan', alpha=0.5)


    # Labels and legend
    ax.set_xlabel("X")
    ax.set_ylabel("Y")
    ax.set_zlabel("Z")
    ax.set_title("3D Motion with Normal Plane")
    axis_limit = 2  # Adjust this value based on your data range
    ax.set_xlim([-axis_limit, axis_limit])
    ax.set_ylim([-axis_limit, axis_limit])
    ax.set_zlim([-axis_limit, axis_limit])
    ax.legend()

    azim = np.arctan2(normal_vector[1], normal_vector[0]) * 180 / np.pi

    # Elevation is calculated based on the z-component of the vector
    plt.gca().set_aspect('equal')
    ax.view_init(elev=90, azim=0)
    ax.grid(False)
    ax.set_xticks([])
    ax.set_yticks([])
    ax.set_zticks([])

    ax.xaxis.set_pane_color((1.0, 1.0, 1.0, 0.0))  # Make panes transparent
    ax.yaxis.set_pane_color((1.0, 1.0, 1.0, 0.0))
    ax.zaxis.set_pane_color((1.0, 1.0, 1.0, 0.0))
    ax.set_axis_off()  # This removes everything including the frame
    plt.show(block=True)
def best_fit_line_3d(points):
    points = np.array(points)
    centroid = np.mean(points, axis=0)
    centered_points = points - centroid
    _, _, Vt = np.linalg.svd(centered_points)
    direction = Vt[0]  # First principal component
    return centroid, direction
def Get3D_path(source_folder):
    AllFiles = []
    for root, dirs, files in os.walk(source_folder):
        for file in files:
            if file.endswith(".csv"):
                input_file_path = os.path.join(root, file)
                AllFiles.append(input_file_path)

    AllFiles = natsorted(AllFiles)
    grouped_data_path = group_files_by_fly(AllFiles)
    return grouped_data_path
def DetermineTiTaMOC(data, fly, trial):
    makeplot = False
    h = 0.05
    trace_start = 200
    start = 250
    end = 1000

    center_points = np.asarray([data["platform-tip_x"].tolist(), data["platform-tip_y"].tolist(), data["platform-tip_z"].tolist()])
    R_mTT_points = np.asarray([data["R-mTT_x"].tolist(), data["R-mTT_y"].tolist(), data["R-mTT_z"].tolist()])
    R_mLT_points = np.asarray([data["R-mLT_x"].tolist(), data["R-mLT_y"].tolist(), data["R-mLT_z"].tolist()])

    L_mTT_points = np.asarray([data["L-mTT_x"].tolist(), data["L-mTT_y"].tolist(), data["L-mTT_z"].tolist()])
    L_mLT_points = np.asarray([data["L-mLT_x"].tolist(), data["L-mLT_y"].tolist(), data["L-mLT_z"].tolist()])

    center_points = np.transpose(center_points)
    R_mTT_points = np.transpose(R_mTT_points)
    R_mLT_points = np.transpose(R_mLT_points)

    L_mTT_points = np.transpose(L_mTT_points)
    L_mLT_points = np.transpose(L_mLT_points)
    print(f"Fly {fly} Trial {trial}")
    contact_count_L = 0
    contact_count_R = 0
    Contact_threshold = 1
    for frame in range(start, end):
        Intersect_R, position_R = check_leg_platform_intersection(R_mTT_points[frame], R_mLT_points[frame], center_points[250:350], center_points[frame], h)
        Intersect_L, position_L = check_leg_platform_intersection(L_mTT_points[frame], L_mLT_points[frame], center_points[250:350], center_points[frame], h)

        if makeplot:
            if frame == 557:
                coords = ReadCoordsAll(data, frame)
                print(Intersect_L, position_L)
                print(frame, "L")
                # plot_motion_vector_with_plane(center_points[250:350], center_points[frame], coords, h)

        if Intersect_L:
            contact_count_L += 1
        else:
            contact_count_L = 0

        if contact_count_L >= Contact_threshold:
            coords = ReadCoordsAll(data, frame)
            print(Intersect_L, position_L)
            print(frame, "L")
            # plot_motion_vector_with_plane(center_points[250:350], center_points[frame], coords, h)
            return frame, "L"

        if Intersect_R:
            contact_count_R += 1
        else:
            contact_count_R = 0

        if contact_count_R >= Contact_threshold:
            coords = ReadCoordsAll(data, frame)
            print(Intersect_R, position_R)
            print(frame, "R")
            # plot_motion_vector_with_plane(center_points[250:350], center_points[frame], coords, h)
            return frame, "R"
    print(0, "NoContact")
    return 0, "NoContact"
def CategorizeTrial(data, fly, trial):
    WingTipDistance = Calculate_segment_length(data, [["L-wing", "R-wing"]])

    NF = detect_Not_Flying(WingTipDistance, 250)
    if NF:
        # Categorized as Not Flying
        return 0, None
    else:
        # Determine if the wing folds
        WingAbTipAngle = Calculate_joint_angle(data, [["R-wing", "R-wing-hinge", "abdomen-tip"], ["L-wing", "L-wing-hinge", "abdomen-tip"]])
        R_mol, RL = detect_moment_of_landing_WingAngle(WingAbTipAngle["R-wing-hinge"], 50, 4, 2)
        L_mol, LL = detect_moment_of_landing_WingAngle(WingAbTipAngle["L-wing-hinge"], 50, 4, 2)
        # mol, L = detect_moment_of_landing(WingTipDistance["L-wing_R-wing"], 2, 4, 3)
        if RL or LL:
            # Categorized as successful landing, may still require further examination for N/A data
            return 1, min([R_mol, L_mol])
        else:
            # Categorized as continuous flying, required further examination for N/A data
            return -1, None
def PlotMOC_platform_distance_distribution(datapath, MOCfile):
    MOCData = Read_Manual_Data(MOCfile)
    Segment = [["R-mLT", "L-platform-tip"], ["R-mLT", "R-platform-tip"], ["R-mLT", "platform-tip"],
               ["R-mTT", "L-platform-tip"], ["R-mTT", "R-platform-tip"], ["R-mTT", "platform-tip"]]
    Collected_segment_data = dict()
    for sg in Segment:
        Collected_segment_data[f"{sg[0]}_{sg[1]}"] = []
    Collected_segment_data["Min"] = []
    for index, r in MOCData.iterrows():
        for i, t in enumerate(r):
            if not isinstance(t, str) and t > -1:
                print(index, i)
                pose_data = pd.read_csv(datapath[f"Fly_{index + 1}"][i])
                segment_data = Calculate_segment_length(pose_data, Segment)
                mins = []
                for sg in Segment:
                    Collected_segment_data[f"{sg[0]}_{sg[1]}"].append(segment_data[f"{sg[0]}_{sg[1]}"][t])
                    mins.append(segment_data[f"{sg[0]}_{sg[1]}"][t])
                Collected_segment_data["Min"].append(min(mins))
    Collected_segment_data = pd.DataFrame(Collected_segment_data)
    sns.stripplot(Collected_segment_data)
    plt.show()

Angles = [["R-mCT", "R-mFT", "R-mTT"], ["L-mCT", "L-mFT", "L-mTT"]]
# Angles = ["R-mFT", "R-mTT", "R-mLT"]
# skeletons = [["R-mTT", "L-platform-tip"], ["R-mTT", "R-platform-tip"], ["R-mTT", "platform-tip"]]
skeletons = [["R-mLT", "L-platform-tip"], ["R-mLT", "R-platform-tip"], ["R-mLT", "platform-tip"]]
skeletons = [["R-mLT", "platform-tip"]]
DataDirPath = r"C:\Users\agrawal-admin\Desktop\TibiaTarsusPlatformODLight-Wayne-2024-10-19\Network-03-14\HCS+_UASKir2.1eGFP\G106-HP1_T2-TiTa"
MOCfilePath = r"C:\Users\agrawal-admin\Desktop\DataFolder\HCS+_UASKir2.1eGFP\G106-HP1_T2-TiTa\2025-03-10\MOCG106.xlsx"
LatencyFilePath = r"C:\Users\agrawal-admin\Desktop\DataFolder\LPAcrossLegsJoints\T2-TiTa\T2-TiTaLP.xlsx"
PredictedLLPath = r"C:\Users\agrawal-admin\Desktop\Landing\LLPrediction_WT.xlsx"
Trials = [f"Trial_{i + 1}" for i in range(20)]


Fly_num = 16
moc_data = pd.read_excel(MOCfilePath)[Trials].iloc[:Fly_num]
LL_data = pd.read_excel(LatencyFilePath)[Trials].iloc[:Fly_num]
predicted_data = pd.read_excel(PredictedLLPath)[Trials].iloc[:Fly_num]
grouped_data_path = Get3D_path(DataDirPath)
print(grouped_data_path)
Target_fly = 7
Target_trial = 2
DetermineTiTaMOC(pd.read_csv(grouped_data_path[f"Fly_{Target_fly}"][Target_trial - 1]), Target_fly, Target_trial)

print(grouped_data_path[f"Fly_{Target_fly}"][Target_trial - 1])

# Predicted_Data = Analyze_3D_pose_data(grouped_data_path, Fly_num)

# OutptuPrediction(Predicted_Data, "LLPrediction_G106")
# for f, k in enumerate(list(grouped_data_path.keys())):
    # MakePlot(grouped_data_path[k], f + 1)
# PlotMOC_platform_distance_distribution(grouped_data_path, MOCfilePath)
error = []


print(moc_data)
collected_joint_angle_L = []
collected_joint_angle_R = []
for index, r in LL_data[Trials].iloc[:Fly_num].iterrows():
    fly_joint_angle_L = []
    fly_joint_angle_R = []
    for i, t in enumerate(r):
        if not isinstance(t, str) and t > -1 and not isinstance(predicted_data.iloc[index][i], str) and t > -1:
            print(f"Fly {index + 1}", f"Trial {i + 1}")
            error.append((t - predicted_data.iloc[index][i] )/ 200)
            # data = pd.read_csv(grouped_data_path[f"Fly_{index + 1}"][i])
            # joints_angle = Calculate_joint_angle(data, Angles)
            # fly_joint_angle_L.append(np.mean(joints_angle["L-mFT"][:250]))
            # fly_joint_angle_R.append(np.mean(joints_angle["R-mFT"][:250]))
        else:
            fly_joint_angle_L.append(np.nan)
            fly_joint_angle_R.append(np.nan)
    collected_joint_angle_L.append(fly_joint_angle_L)
    collected_joint_angle_R.append(fly_joint_angle_R)

sns.stripplot(error)
plt.ylabel("Seconds (s)")
plt.xlabel("Trial manual latency - prediction")
plt.show()

fig, ax = plt.subplots()
sns.lineplot(collected_joint_angle_R, legend=False)
ax.set_xlabel("Trial", fontsize=25)
ax.set_ylabel("R FT joint angle", fontsize=25)
plt.xticks([1, 10, 20], fontsize=25)
plt.yticks([0, 45, 90], fontsize=25)
for spine in ax.spines.values():
    spine.set_linewidth(2)
plt.savefig("FT angle R")
plt.tight_layout()
# plt.show()


fig, ax = plt.subplots()
sns.lineplot(collected_joint_angle_L, legend=False)
ax.set_xlabel("Trial", fontsize=25)
ax.set_ylabel("L FT joint angle", fontsize=25)
plt.xticks([1, 10, 20], fontsize=25)
plt.yticks([0, 45, 90], fontsize=25)
for spine in ax.spines.values():
    spine.set_linewidth(2)
plt.savefig("FT angle L")
plt.tight_layout()
# plt.show()


"""

collected_joint_angle_L = pd.DataFrame(collected_joint_angle_L)
collected_joint_angle_R = pd.DataFrame(collected_joint_angle_R)


mean_values_R = collected_joint_angle_R.mean(axis=0, skipna=True)  # Compute mean for each trial (column)
sem_values_R = collected_joint_angle_R.sem(axis=0, skipna=True)  # Compute SEM for each trial (column)

mean_values_L = collected_joint_angle_L.mean(axis=0, skipna=True)  # Compute mean for each trial (column)
sem_values_L = collected_joint_angle_L.sem(axis=0, skipna=True)

fig, ax = plt.subplots()

sns.lineplot(x=mean_values_R.index + 1, y=mean_values_R,
             errorbar=("ci", 0),  # Disables default CI
             ax=ax, label="Mean", color="Navy")

sns.lineplot(x=mean_values_L.index + 1, y=mean_values_L,
             errorbar=("ci", 0),  # Disables default CI
             ax=ax, label="Mean", color="skyblue")

# Add error bars manually
ax.errorbar(mean_values_R.index + 1, mean_values_R, yerr=sem_values_R, fmt='o', color="Navy", capsize=5)
ax.errorbar(mean_values_L.index + 1, mean_values_L, yerr=sem_values_R, fmt='o', color="skyblue", capsize=5)

# sns.lineplot(collected_joint_angle_R, legend=False)
ax.set_xlabel("Trial", fontsize=25)
ax.set_ylabel("FT joint angle", fontsize=25)
plt.xticks([1, 10, 20], fontsize=25)
plt.yticks([0, 45, 90], fontsize=25)
plt.tick_params(axis="both", width=3, length=10)
for spine in ax.spines.values():
    spine.set_linewidth(2)
plt.savefig("FT angle mean")
plt.tight_layout()
plt.show()



h = 0.05
moc_error = []
Contact_relative_position = []
Latency = []
Angles_at_MOC = []
for index, r in LL_data[Trials].iloc[:Fly_num].iterrows():
    for i, t in enumerate(r):
        if not isinstance(t, str) and t > -1:
            print(f"Fly {index + 1}", f"Trial {i + 1}")
            data = pd.read_csv(grouped_data_path[f"Fly_{index + 1}"][i])
            start = 200
            end = 800

            center_points = np.asarray(
                [data["platform-tip_x"].tolist(), data["platform-tip_y"].tolist(), data["platform-tip_z"].tolist()])
            R_mTT_points = np.asarray([data["R-mTT_x"].tolist(), data["R-mTT_y"].tolist(), data["R-mTT_z"].tolist()])
            R_mLT_points = np.asarray([data["R-mLT_x"].tolist(), data["R-mLT_y"].tolist(), data["R-mLT_z"].tolist()])

            center_points = np.transpose(center_points)
            R_mTT_points = np.transpose(R_mTT_points)
            R_mLT_points = np.transpose(R_mLT_points)

            for frame in range(start, end):
                Intersect, position = check_leg_platform_intersection(R_mTT_points[frame], R_mLT_points[frame], center_points[250:350], center_points[frame], h)
                if Intersect:
                    name_width = 8
                    num_width = 3
                    name = "Prediction:"
                    name1 = "Manual:"
                    name2 = "Error:"
                    # print(f"{name:<{name_width}} {str(frame):>{num_width}} {name1:<{name_width}} {str(t):>{num_width}} {name2:<{name_width}} {str(frame - t):>{num_width}}")
                    if (LL_data.loc[index][i] / 250) < 1:
                        moc_error.append(frame - t)
                        Contact_relative_position.append(position)
                        Latency.append(LL_data.loc[index][i] / 250)
                    #Angles_at_MOC.append(calculate_angle(data[f"{Angles[0]}_x"][frame], data[f"{Angles[0]}_y"][frame], data[f"{Angles[0]}_z"][frame],
                    #                                     data[f"{Angles[1]}_x"][frame], data[f"{Angles[1]}_y"][frame], data[f"{Angles[1]}_z"][frame],
                     #                                    data[f"{Angles[2]}_x"][frame], data[f"{Angles[2]}_y"][frame], data[f"{Angles[2]}_z"][frame]))
                    if abs(frame - t) >= 20:
                        print("Large error occured")
                        coords = ReadCoordsAll(data, frame)
                        # plot_motion_vector_with_plane(center_points[250:350], center_points[frame], coords, h)
                    break

df = {
    "ContactPoints": Contact_relative_position,
    "Latency": Latency
}

df = pd.DataFrame(df)
fig, ax = plt.subplots()

sns.scatterplot(x="ContactPoints", y="Latency", data=df, markers="o", s=100, alpha=0.8, legend=False)

ax.set_xlabel("Ti-Ta relative contact point", fontsize=25)
ax.set_ylabel("Latency (s)", fontsize=25)
plt.xticks([0, 1], fontsize=25)
plt.yticks([0, 1], fontsize=25)
for spine in ax.spines.values():
    spine.set_linewidth(2)
plt.tick_params(axis="both", width=3, length=10)
plt.title("G106", fontsize=25)
plt.ylim([-0.1, 1.1])
plt.xlim([-0.1, 1.1])
plt.tight_layout()
plt.show()

"""



