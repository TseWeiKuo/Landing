import os
import numpy as np
import pandas as pd
from openpyxl.styles import PatternFill
from openpyxl import load_workbook
from scipy.signal import find_peaks, hilbert
import matplotlib.pyplot as plt


from kinematic_object import Group, Trial
import warnings
from sklearn.utils import resample
import seaborn as sns
import pywt
warnings.filterwarnings(action="ignore", category=FutureWarning)
ax_flip_count = 0
total_count = 0
"""
These functions are responsible for preprocessing of angle data and 3D pose data
"""
class SimpleCalculation:
    def calculate_mean_diff(self, data1, data2):
        return np.mean(data1) - np.mean(data2)

    def calculate_median_diff(self, data1, data2):
        return np.median(data1) - np.median(data2)
    # Smooth the input data using ema
    def exponential_moving_average(self, data, alpha):
        if isinstance(data, pd.Series):
            data = data.tolist()
        smoothed_data = [data[0]]
        for i in range(1, len(data)):
            smoothed_data.append(alpha * data[i] + (1 - alpha) * smoothed_data[-1])
        return smoothed_data
    # Normalized the input list based on min-max method
    def normalize_list(self, data, method="min-max"):
        if method == "min-max":
            min_val = min(data)
            max_val = max(data)
            if max_val == min_val:
                raise ValueError("Cannot perform Min-Max normalization when all values are the same.")
            return [(x - min_val) / (max_val - min_val) for x in data]
        if method == "z-score":
            """
            Perform Z-score normalization on a 1D signal.

            Parameters:
            -----------
            signal : array-like
                Input signal (list or 1D numpy array).

            Returns:
            --------
            normalized_signal : numpy.ndarray
                Z-score normalized signal with mean = 0 and std = 1.
            """
            signal = np.asarray(data)
            mean = np.mean(signal)
            std = np.std(signal)
            if std == 0:
                return np.zeros_like(signal)  # Avoid division by zero
            return (signal - mean) / std
    # Calculate the distance between two point in 3D space
    def Calculate_distance_between_points(self, x, y, z, x1, y1, z1):
        return np.sqrt((x - x1) ** 2 + (y - y1) ** 2 + (z - z1) ** 2)
    # Calculate the angle between pt1, pt2, and pt3 in 3D space
    def calculate_angle(self, x1, y1, z1, x2, y2, z2, x3, y3, z3):
        # Define points as numpy arrays
        pt1 = np.array([x1, y1, z1])
        pt2 = np.array([x2, y2, z2])
        pt3 = np.array([x3, y3, z3])

        # Define vectors
        vecA = pt1 - pt2
        vecB = pt3 - pt2

        # Calculate dot product and magnitudes
        dot_product = np.dot(vecA, vecB)
        magnitude_A = np.linalg.norm(vecA)
        magnitude_B = np.linalg.norm(vecB)

        # Calculate angle in radians
        angle_rad = np.arccos(dot_product / (magnitude_A * magnitude_B))

        # Convert to degrees if needed
        angle_deg = np.degrees(angle_rad)

        return angle_deg
    # Transpose a dataframe
    def TransposeData(self, df):
        df = pd.DataFrame(df)
        df = df.T
        return df
    # Calculate the specified segments' length and return a dictionary
    def Calculate_segment_length(self, trial_info:Trial, skeletons):
        # Initialize the dictionary data to stor segments' length
        collected_seg_length_data = dict()

        # Iterate through the specified segments
        for seg in skeletons:

            # If segment not in dictionary, create one
            if f"{seg[0]}_{seg[1]}" not in collected_seg_length_data.keys():
                collected_seg_length_data[f"{seg[0]}_{seg[1]}"] = []

            for f in range(trial_info.total_frames_number):
                collected_seg_length_data[f"{seg[0]}_{seg[1]}"].append(
                    self.Calculate_distance_between_points(
                        x=trial_info.trial_data[f"{seg[0]}"].x_coord[f],
                        y=trial_info.trial_data[f"{seg[0]}"].y_coord[f],
                        z=trial_info.trial_data[f"{seg[0]}"].z_coord[f],
                        x1=trial_info.trial_data[f"{seg[1]}"].x_coord[f],
                        y1=trial_info.trial_data[f"{seg[1]}"].y_coord[f],
                        z1=trial_info.trial_data[f"{seg[1]}"].z_coord[f]))

        # Convert the dictionary to dataframe
        # collected_seg_length_data = pd.DataFrame(collected_seg_length_data)

        # return the data
        return collected_seg_length_data
    # Calculate the angles between specified key points.
    def Calculate_joint_angle(self, data:Trial, angles):

        collected_angle_data = dict()

        for ag in angles:
            if f"{ag[1]}" not in collected_angle_data.keys():
                collected_angle_data[f"{ag[1]}"] = []

            for f in range(data.total_frames_number):
                angle = self.calculate_angle(x1=data.trial_data[f"{ag[0]}"].x_coord[f],
                                             y1=data.trial_data[f"{ag[0]}"].y_coord[f],
                                             z1=data.trial_data[f"{ag[0]}"].z_coord[f],
                                             x2=data.trial_data[f"{ag[1]}"].x_coord[f],
                                             y2=data.trial_data[f"{ag[1]}"].y_coord[f],
                                             z2=data.trial_data[f"{ag[1]}"].z_coord[f],
                                             x3=data.trial_data[f"{ag[2]}"].x_coord[f],
                                             y3=data.trial_data[f"{ag[2]}"].y_coord[f],
                                             z3=data.trial_data[f"{ag[2]}"].z_coord[f])
                collected_angle_data[f"{ag[1]}"].append(angle)
            collected_angle_data[f"{ag[1]}"] = np.array(collected_angle_data[f"{ag[1]}"])
        return collected_angle_data
    # Calculate the absolute value of derivative of the input data
    def Calculate_derivative(self, data):
        if len(data) < 2:
            raise ValueError("The input list must contain at least two elements to calculate a derivative.")

        # Compute the differences between consecutive elements
        data = [data[i + 1] - data[i] for i in range(len(data) - 1)]
        data = [abs(x) for x in data]
        return data
    # Calculate the angle between a vector and a normal plane given the plane's normal vector
    def angle_between_vectors(self, v, n):
        """
        Calculate the angle (in degrees) between a vector and a plane
        given the vector and the normal vector of the plane.

        Parameters:
        v : array-like
            The input vector.
        n : array-like
            The normal vector of the plane.

        Returns:
        angle_deg : float
            The angle between the vector and the plane in degrees.
        """
        v = np.array(v)
        n = np.array(n)

        # Normalize the vectors
        v_norm = np.linalg.norm(v)
        n_norm = np.linalg.norm(n)

        # Dot product
        dot_product = np.dot(v, n)

        # Angle between v and n
        phi_rad = np.arccos(np.clip(dot_product / (v_norm * n_norm), -1.0, 1.0))
        # Angle between vector and plane is 90 - phi
        theta_rad = np.pi / 2 - phi_rad
        # Convert to degrees
        angle_deg = np.degrees(theta_rad)

        return angle_deg
    # Calculate the intersection between a segment and a plane.
    def line_plane_intersection(self, p1, p2, normal, d):
        # Line represented as p(t) = p1 + t * (p2 - p1)
        direction = np.array(p2) - np.array(p1)
        denom = np.dot(normal, direction)

        if abs(denom) < 1e-6:  # Parallel case
            return None, None

        t = -(np.dot(normal, p1) + d) / denom

        if 0 <= t <= 1:  # Ensure intersection is within the segment
            return p1 + t * direction, t
        return None, None
    # Determine if the intersection is inside a circle determined by the radius and certain point
    def is_inside_circle(self, intersection, center, radius):
        distance_to_tip = np.linalg.norm(intersection - center)
        return np.linalg.norm(intersection - center) <= radius
    # Finding the direction of the best fit line in 3D space
    def best_fit_line_3d(self, points):
        points = np.array(points)
        centroid = np.mean(points, axis=0)
        centered_points = points - centroid
        _, _, Vt = np.linalg.svd(centered_points)
        direction = Vt[0]  # First principal component
        return centroid, direction
    # Detect the contact between leg segment and side of cylindrical platform
    def check_cylinder_side_intersection(self, A, B, P1, d, r, h, n_steps=100):
        """
        Check if the segment AB intersects the side surface of a finite cylinder,
        using the cylinder's top point instead of the base.

        Parameters:
            A, B : ndarray
                Endpoints of the leg segment (3D).
            P1 : ndarray
                A point at the top of the cylinder axis.
            d : ndarray
                Unit vector of the cylinder axis (pointing from base to top).
            r : float
                Radius of the cylinder.
            h : float
                Height of the cylinder.
            n_steps : int
                Number of steps to interpolate along the segment.

        Returns:
            (bool, point, min_dist): True and intersection point if intersects, otherwise False, None
        """

        A, B, P1, d = map(np.asarray, (A, B, P1, d))
        d = d / np.linalg.norm(d)  # ensure it's a unit vector

        # Calculate the base of the cylinder from the top point
        P0 = P1 - h * d
        min_dist = np.inf
        for t in np.linspace(0, 1, n_steps):
            Pt = A + t * (B - A)
            v = Pt - P0
            proj_len = np.dot(v, d)
            dist_to_axis = np.linalg.norm(v - proj_len * d)
            if dist_to_axis < min_dist:
                min_dist = dist_to_axis
            if 0 <= proj_len <= h:
                if np.isclose(dist_to_axis, r) or dist_to_axis < r:
                    return True, Pt, min_dist
        return False, None, min_dist
    def ReadAndTranspose(self, point, kinematic_data: Trial):
        return np.transpose(np.asarray([kinematic_data.trial_data[point].x_coord,
                                        kinematic_data.trial_data[point].y_coord,
                                        kinematic_data.trial_data[point].z_coord]))
    def calculate_platform_surfaces(self, platform_traces, platform_center, platform_offset, radius, height):
        global ax_flip_count
        global total_count
        total_count += 1
        centroid, direction = self.best_fit_line_3d(platform_traces)
        motion_vec = np.asarray(platform_traces[-1]) - np.asarray(platform_traces[0])
        # print(np.dot(direction, motion_vec))
        if np.dot(direction, motion_vec) < 0:
            direction = -direction
        # Generate points along the best-fit line
        t_vals = np.linspace(-10, 10, 100)
        line_points = centroid + np.outer(t_vals, direction)

        # Generate a normal plane to the direction vector at the centroid
        normal_vector = direction
        # Create two perpendicular vectors using cross products
        if np.allclose(normal_vector, [1, 0, 0]):  # Handle edge case if aligned with x-axis
            perp_vector1 = np.cross(normal_vector, [0, 1, 0])
        else:
            perp_vector1 = np.cross(normal_vector, [1, 0, 0])

        perp_vector1 /= np.linalg.norm(perp_vector1)  # Normalize
        perp_vector2 = np.cross(normal_vector, perp_vector1)
        perp_vector2 /= np.linalg.norm(perp_vector2)  # Normalize

        # Generate a circular grid
        u_vals = np.linspace(-radius, radius, 50)
        v_vals = np.linspace(-radius, radius, 50)
        U, V = np.meshgrid(u_vals, v_vals)

        # Convert to polar coordinates to filter for a circular region
        mask = U ** 2 + V ** 2 <= radius ** 2  # Boolean mask for circle

        # Apply mask to keep only circular region
        U = U[mask]
        V = V[mask]

        platform_plane_origin = platform_center + platform_offset * direction  # Shift the plane up
        plane_points = platform_plane_origin + U[..., None] * perp_vector1 + V[..., None] * perp_vector2

        # Define top and bottom centers of the cylinder
        cylinder_top = platform_plane_origin
        cylinder_bottom = cylinder_top - direction * height
        # Generate points around the circumference

        theta = np.linspace(0, 2 * np.pi, 60)
        circle_top = np.array([
            cylinder_top + radius * (np.cos(t) * perp_vector1 + np.sin(t) * perp_vector2)
            for t in theta
        ])
        circle_bottom = np.array([
            cylinder_bottom + radius * (np.cos(t) * perp_vector1 + np.sin(t) * perp_vector2)
            for t in theta
        ])
        # Create side surface polygons (vertical quads)
        verts = []
        for i in range(len(theta) - 1):
            quad = [circle_bottom[i], circle_bottom[i + 1], circle_top[i + 1], circle_top[i]]
            verts.append(quad)

        return line_points, plane_points, verts, cylinder_top, cylinder_bottom, direction, perp_vector1, perp_vector2
    def angle_between_vectors_360(self, a, b, degrees=True):
        """
        Calculate the angle between two vectors a and b.

        Parameters:
        -----------
        a, b : array-like
            Input vectors (1D arrays).
        degrees : bool
            If True, return angle in degrees. Otherwise, radians.

        Returns:
        --------
        angle : float
            Angle between vectors in radians or degrees.
        """
        a = np.array(a)
        b = np.array(b)

        # Normalize both vectors
        a_norm = a / np.linalg.norm(a)
        b_norm = b / np.linalg.norm(b)

        # Calculate angle using atan2 of the determinant and dot product
        dot = np.dot(a_norm, b_norm)
        det = a_norm[0] * b_norm[1] - a_norm[1] * b_norm[0]  # 2D cross product (scalar)

        angle_rad = np.arctan2(det, dot)
        angle_deg = np.degrees(angle_rad)

        # Convert from [-180, 180] to [0, 360]
        return angle_deg % 360
    def angle_between_vectors_unsigned(self, a, b, degrees=True):
        a = np.array(a)
        b = np.array(b)

        # Normalize both vectors
        a_norm = a / np.linalg.norm(a)
        b_norm = b / np.linalg.norm(b)

        # Clamp the dot product to avoid floating-point errors
        dot = np.clip(np.dot(a_norm, b_norm), -1.0, 1.0)

        angle_rad = np.arccos(dot)

        return np.degrees(angle_rad) if degrees else angle_rad
    def Coordinates_transformation(self, trial_info:Trial, new_axis):
        for p in trial_info.trial_data.keys():
            points = np.array([trial_info.trial_data[p].x_coord, trial_info.trial_data[p].y_coord, trial_info.trial_data[p].z_coord]).T
            transformed_points = points @ new_axis
            transformed_points = transformed_points.T
            trial_info.trial_data[p].x_coord = transformed_points[0]
            trial_info.trial_data[p].y_coord = transformed_points[1]
            trial_info.trial_data[p].z_coord = transformed_points[2]
    def transform_coords_and_calculate_platform_data(self, trial_info:Trial, platform_offset, radius, platform_height, frame=0):

        if frame == 0:
            start = int(trial_info.moc)
        else:
            start = frame

        center_points = self.ReadAndTranspose("platform-tip", trial_info)
        platform_ctr_pts_traces = np.array(center_points[200:250])

        line_points_before, plane_points_before, verts_before, cylinder_top, cylinder_bottom, direction_before, perp_vector1_before, perp_vector2_before = (
            self.calculate_platform_surfaces(platform_traces=platform_ctr_pts_traces,
                                             platform_center=center_points[start],
                                             platform_offset=platform_offset,
                                             radius=radius,
                                             height=platform_height))

        """# coordinates transformation
        v1 = perp_vector1_before  # new Y-axis
        v2_proj = direction_before - np.dot(direction_before, v1) * v1  # remove v1 component from v2
        v2 = v2_proj / np.linalg.norm(v2_proj)  # new Z-axis (orthogonalized)
        v0 = np.cross(v1, v2)  # new X-axis, orthogonal to both
        v0 = v0 / np.linalg.norm(v0)
        # Double-check orthogonality (optional)
        assert np.isclose(np.dot(v0, v1), 0)
        assert np.isclose(np.dot(v0, v2), 0)
        assert np.isclose(np.dot(v1, v2), 0)

        # Each column is a basis vector
        R = np.column_stack([v0, v1, v2])  # shape (3, 3)

        # Transform coordinates based on platform motion and L-R-mBC vector
        self.Coordinates_transformation(trial_info, R)

        center_points = self.ReadAndTranspose("platform-tip", trial_info)
        platform_ctr_pts_traces = np.array(center_points[200:250])

        line_points_after, plane_points_after, verts_after, cylinder_top, cylinder_bottom, direction_after, perp_vector1_after, perp_vector2_after = (
            self.calculate_platform_surfaces(platform_traces=platform_ctr_pts_traces,
                                             platform_center=center_points[start],
                                             platform_offset=platform_offset,
                                             radius=radius,
                                             height=platform_height))"""

        return line_points_before, plane_points_before, verts_before, cylinder_top, cylinder_bottom, direction_before, perp_vector1_before, perp_vector2_before
        # return line_points_after, plane_points_after, verts_after, cylinder_top, cylinder_bottom, direction_after, perp_vector1_after, perp_vector2_after
    def calculate_rhythmicity(self, signal, fps):


        analytic_signal = hilbert(signal)
        phase = np.angle(analytic_signal)

        # 3. Compute pACF (Phase autocorrelation)
        max_lag_cycles = int(len(signal) / (fps / 10)) # up to 10 cycles
        lags = np.arange(0, max_lag_cycles, 0.1)  # lags in cycles
        pACF = []
        samples_per_cycle = int(fps / 10)
        # print(max_lag_cycles)
        for lag in lags:
            shift = int(lag * samples_per_cycle)  # convert lag in cycles → samples
            if shift == 0:
                pACF.append(1.0)  # autocorrelation at lag 0 = 1
                continue
            if shift != len(signal):
                phase_shifted = np.roll(phase, -shift)

                phase_diff = phase[:-shift] - phase[shift:]
                # phase_diff = phase - phase_shifted
                # plt.plot(phase)
                # plt.plot(phase_shifted)
                # plt.show()
                # compute PLV = |average of complex exponentials|
                plv = np.abs(np.mean(np.exp(1j * phase_diff)))
                # print(plv)
                pACF.append(plv)

        pACF = np.array(pACF)
        # pACF = self.normalize_list(pACF)
        # 4. Normalize pACF to explained variance function
        pACF_norm = pACF / np.sum(pACF)

        # 5. Cumulative function
        CF = np.cumsum(pACF_norm)
        # 6. Lifetime score = first lag where CF > 0.9
        threshold = 0.8
        lifetime_idx = np.where(CF >= threshold)[0][0]
        lifetime_cycles = lags[lifetime_idx]

        MakePlot = True
        print(lifetime_cycles / max_lag_cycles)
        if MakePlot:
            # 7. Plot
            plt.figure(figsize=(12, 4))
            plt.subplot(1, 3, 1)
            plt.plot(lags, pACF, 'o-')
            plt.xlabel("Lag (cycles)")
            plt.ylabel("pACF")
            plt.title("Phase Autocorrelation")
            plt.ylim(-0.1, 1.1)
            plt.yticks([0, 1])

            plt.subplot(1, 3, 2)
            plt.plot(lags, pACF_norm, 'o-')
            plt.xlabel("Lag (cycles)")
            plt.title("Normalized pACF")

            plt.subplot(1, 3, 3)
            plt.plot(lags, CF, 'o-')
            plt.axhline(threshold, color='r', linestyle='--')
            plt.axvline(lifetime_cycles, color='g', linestyle='--')
            plt.xlabel("Lag (cycles)")
            plt.ylabel("Cumulative pACF")
            # plt.title(f"Lifetime = {lifetime_cycles / max_lag_cycles:.2f} cycles")
            plt.show()

        # return lifetime_cycles / max_lag_cycles
    def calculate_velocity(self, trial_info:Trial, points):
        velocity_data = dict()
        for p in points:
            dt = 1 / trial_info.fps  # fps = frames per second
            # Stack coordinates
            coords = np.stack([trial_info.trial_data[p].x_coord, trial_info.trial_data[p].y_coord, trial_info.trial_data[p].z_coord], axis=1)
            # Compute velocity using finite differences
            velocities = np.diff(coords, axis=0) / dt  # shape (n_timepoints-1, 3)
            # Compute speed (magnitude)
            speed = np.linalg.norm(velocities, axis=1)  # shape (n_timepoints-1,)
            velocity_data[p] = speed
        return velocity_data
    def detect_angle_deviation(self, angle_trace, MOC, baseline_window=50, threshold_sd=2, consecutive_frames=3):

        start = max(MOC - baseline_window, 0)
        end = MOC
        baseline = angle_trace[start:end]
        baseline_mean = np.mean(baseline)
        baseline_sd = np.std(baseline)


        deviation_zscore = (angle_trace - baseline_mean) / baseline_sd


        above_threshold = np.abs(deviation_zscore) > threshold_sd

        count = 0
        deviation_frame = None
        for i in range(MOC, len(angle_trace)):
            if above_threshold[i]:
                count += 1
                if count >= consecutive_frames:
                    deviation_frame = i - consecutive_frames + 1  # first frame of deviation
                    break
            else:
                count = 0  # reset counter if not consecutive

        return deviation_frame, deviation_zscore
    def determine_side_contact(self, trace, dist_threshold=0.5, vel_threshold=0.02):
        stable_count = 0
        velocity = abs(np.gradient(trace))
        Contact = False
        MOC = np.nan
        for i in range(len(trace)):
            if trace[i] <= dist_threshold and velocity[i] <= vel_threshold:
                stable_count += 1
            else:
                stable_count = 0
            if stable_count >= 3:
                Contact = True
                MOC = i - 3
        return Contact, MOC

    def Bootstrapping_test(self, data1, data2, n_samps):

        num_bootstrap_samples = n_samps
        original_mean_diff = self.calculate_mean_diff(data1, data2)

        # Bootstrap resampling
        bootstrap_mean_diffs = []
        resample_data = np.concatenate((data1, data2))
        for _ in range(num_bootstrap_samples):
            # Resample with replacement
            bootstrap_sample1 = resample(resample_data, n_samples=len(data1))
            bootstrap_sample2 = resample(resample_data, n_samples=len(data2))
            bootstrap_mean_diff = self.calculate_mean_diff(bootstrap_sample1, bootstrap_sample2)


            bootstrap_mean_diffs.append(bootstrap_mean_diff)

        k = 0
        for m in bootstrap_mean_diffs:
            if abs(m) > abs(original_mean_diff):
                k += 1
        Mean_diff_p_value = (np.sum(np.abs(bootstrap_mean_diffs) >= np.abs(original_mean_diff))) / (
            num_bootstrap_samples)


        return Mean_diff_p_value
    def smoothing(self, data, window=5, polyorder=3):
        from scipy.signal import savgol_filter
        return savgol_filter(data, window_length=window, polyorder=polyorder)
    def Normalized_time(self, data, length=250):
        from scipy.interpolate import interp1d
        x_old = np.linspace(0, 1, len(data))
        x_new = np.linspace(0, 1, length)
        f = interp1d(x_old, data, kind='linear')
        signal = f(x_new)
        return signal

    def projected_signed_angle(self, tibia_pt, tip_pt, pt1, pt2, plane_normal):
        """
        Compute the CLOCKWISE signed angle from pt1->pt2 (x-axis)
        to tibia->tip vector, after projecting both onto a plane.

        Parameters
        ----------
        tibia_pt, tip_pt : (3,)
            Defines leg vector.
        pt1, pt2 : (3,)
            Defines reference x-axis.
        plane_normal : (3,)
            Normal of the plane to project onto (e.g. platform direction).
        degrees : bool
            Return degrees if True, else radians.

        Returns
        -------
        angle : float
            Clockwise angle from reference axis to leg vector.
            Range: (-180, 180]
        """

        tibia_pt = np.asarray(tibia_pt, float)
        tip_pt = np.asarray(tip_pt, float)
        pt1 = np.asarray(pt1, float)
        pt2 = np.asarray(pt2, float)
        n = np.asarray(plane_normal, float)

        # Normalize normal
        n_norm = np.linalg.norm(n)
        if n_norm < 1e-8:
            return np.nan
        n = n / n_norm

        # Define vectors
        x_axis = pt2 - pt1
        v_leg = tip_pt - tibia_pt

        # Project onto plane
        x_proj = x_axis - np.dot(x_axis, n) * n
        v_proj = v_leg - np.dot(v_leg, n) * n

        x_norm = np.linalg.norm(x_proj)
        v_norm = np.linalg.norm(v_proj)

        if x_norm < 1e-8 or v_norm < 1e-8:
            return np.nan

        x_unit = x_proj / x_norm
        v_unit = v_proj / v_norm

        # CCW signed angle
        cross_term = np.dot(n, np.cross(x_unit, v_unit))
        dot_term = np.clip(np.dot(x_unit, v_unit), -1.0, 1.0)

        angle_ccw = np.arctan2(cross_term, dot_term)

        # Convert to clockwise
        angle_cw = -angle_ccw

        angle_cw = np.degrees(angle_cw)

        return angle_cw



    def get_stable_point(self, trial_info:Trial, point, start, average_frames=0):
        x = np.mean(trial_info.trial_data[point].x_coord[start - average_frames:start])
        y = np.mean(trial_info.trial_data[point].y_coord[start - average_frames:start])
        z = np.mean(trial_info.trial_data[point].z_coord[start - average_frames:start])
        return np.array([x, y, z])
"""
These functions are responsible for detecting various characteristic of the angle and 3D data
"""
class DetectCharacteristics:
    def __init__(self, radius=0, FPS=0):
        self.radius = radius
        self.calculator = SimpleCalculation()
        self.fps = FPS

    def check_leg_platform_intersection(self, leg_p1, leg_p2, direction, center_point, platform_offset):
        # Compute platform radius
        # Compute the plane equation

        platform_plane_origin = center_point + platform_offset * direction  # Shift the plane up
        d = -np.dot(direction, platform_plane_origin)

        # Find intersection point
        intersection, intersect_proportion = self.calculator.line_plane_intersection(np.array(leg_p1), np.array(leg_p2), direction, d)
        # print(f"Top intersection: {intersection} proportion:{intersect_proportion}")
        if intersection is not None and self.calculator.is_inside_circle(intersection, np.array(platform_plane_origin), self.radius):
            return True, intersect_proportion
        return False, None
    def detect_leg_stuck(self, CT_FT_ag, CT_TT_ag, CT_LT_ag, moc,  mol, alpha):
        wind_size = 3

        smoothed_CT_FT = self.calculator.exponential_moving_average(CT_FT_ag[moc:mol], alpha)
        smoothed_CT_TT = self.calculator.exponential_moving_average(CT_TT_ag[moc:mol], alpha)
        smoothed_CT_LT = self.calculator.exponential_moving_average(CT_LT_ag[moc:mol], alpha)

        from scipy.stats import linregress

        for i in range(wind_size, len(smoothed_CT_FT), wind_size):
            w1 = smoothed_CT_FT[i:i + wind_size]
            w2 = smoothed_CT_TT[i:i + wind_size]
            w3 = smoothed_CT_LT[i:i + wind_size]

            TT_angle_threshold = 0
            LT_angle_threshold = 0

            # Condition 1: mean of list1 window is 0
            cond1 = np.mean(w1) > 0.0

            df = None

            if i < 50:
                x = np.arange(i)  # x values for regression
                slope2, _, _, _, _ = linregress(x, smoothed_CT_TT[:i])
                slope3, _, _, _, _ = linregress(x, smoothed_CT_LT[:i])
                df = pd.DataFrame({
                    'm2': smoothed_CT_TT[:i],
                    'm3': smoothed_CT_LT[:i]
                })

                # Compute correlation matrix
                corr = df.corr(method="pearson")

            else:
                x = np.arange(50)  # x values for regression
                slope2, _, _, _, _ = linregress(x, smoothed_CT_TT[i - 50:i])
                slope3, _, _, _, _ = linregress(x, smoothed_CT_LT[i - 50:i])

                df = pd.DataFrame({
                    'm2': smoothed_CT_TT[i - 50:i],
                    'm3': smoothed_CT_LT[i - 50:i]
                })

                # Compute correlation matrix
                corr = df.corr(method="pearson")


            cond2 = slope2 > 0
            cond3 = slope3 > 0

            cond4 = np.mean(w2) > TT_angle_threshold and np.mean(w3) > LT_angle_threshold
            # cond5 = corr["m2"]["m3"] > 0.8
            if cond1 and cond2 and cond4 and cond3:
                return i - 1  # return end index of the matching window
        return 0
    def detect_stable_posture(self, Angle_data):
        sum_move = 0
        for k in Angle_data.keys():
            postureData = self.calculator.Calculate_derivative(self.calculator.exponential_moving_average(Angle_data[k][:self.fps], 0.1))
            peaks, _ = find_peaks(postureData, height=3)
            sum_move += len(peaks)
        print(sum_move)
        if sum_move > 10:
            return False
        else:
            return True
    def find_first_trough_CT_ang(self, Angle_data):
        from scipy.stats import linregress
        wind = 5
        for f in range(len(Angle_data) - wind):
            x = np.arange(wind)
            slope, _, _, _, _ = linregress(x, Angle_data[f:f + wind])
            if slope >= 5:
                return f
        return 0
    def find_last_trough_CT_ang(self, Angle_data):
        counts = 0
        derivative = np.gradient(Angle_data)
        for d in range(len(derivative)):
            if derivative[d] < -4:
                counts += 1
            if counts >= 5 and d == len(derivative) - 1:
                return d
            if abs(derivative[d]) < 2 and counts >= 5:
                return d
        return 0
    def detect_cycle_present(self, Angle_data):
        derivative = np.gradient(Angle_data)
        rise = False
        drops = 0
        sequence = []
        for d in range(len(derivative)):
            if derivative[d] > 6 and not rise:
                rise = True
                sequence.append(d)
            if derivative[d] < -4:
                drops += 1
            if drops > 5 and d == len(derivative) - 1:
                sequence.append(d)
                break
            if abs(derivative[d]) < 2 and drops > 0:
                sequence.append(d)
                break

        return sequence
    def ReadCoordsAll(self, kinematic_data:Trial, fnum):
        points = ["L-wing", "L-wing-hinge", "R-wing", "R-wing-hinge", "abdomen-tip", "platform-tip", "L-platform-tip",
                  "R-platform-tip", "platform-axis",
                  "R-fBC", "R-fCT", "R-fFT", "R-fTT", "R-fLT",
                  "R-mBC", "R-mCT", "R-mFT", "R-mTT", "R-mLT",
                  "R-hBC", "R-hCT", "R-hFT", "R-hTT", "R-hLT",
                  "L-fBC", "L-fCT", "L-fFT", "L-fTT", "L-fLT",
                  "L-mBC", "L-mCT", "L-mFT", "L-mTT", "L-mLT",
                  "L-hBC", "L-hCT", "L-hFT", "L-hTT", "L-hLT"]
        coords = dict()
        for p in points:
            coords[p] = np.asarray([kinematic_data.trial_data[p].x_coord[fnum],
                                    kinematic_data.trial_data[p].y_coord[fnum],
                                    kinematic_data.trial_data[p].z_coord[fnum]])
        return coords
    def leg_search_cycle_start_classifier(self, peaks, troughs, angle_data):
        peaks_num = len(peaks)
        troughs_num = len(troughs)
        if troughs_num == 0 and peaks_num == 1:
            # print("need one more troughs")
            last_troughs = self.find_last_trough_CT_ang(angle_data)
            if last_troughs > 0:
                troughs = np.insert(troughs, 0, self.find_first_trough_CT_ang(angle_data))
                troughs = np.insert(troughs, 1, last_troughs)
        if troughs_num > 0 and abs(troughs_num - peaks_num) < 2:
            # print("Troughs detected")
            # print(troughs)
            if troughs[0] > 10:
                initial_troughs = self.find_first_trough_CT_ang(angle_data[:troughs[0]])
                if initial_troughs > 0:
                    troughs = np.insert(troughs, 0, initial_troughs)

            if self.all_elements_in_between(troughs, peaks):
                # print("need one more troughs")
                last_troughs = self.find_last_trough_CT_ang(angle_data)
                if last_troughs > 0:
                    troughs = np.insert(troughs, 1, last_troughs)
                    # if not self.all_elements_in_between(peaks[-1], troughs):
                    #     troughs = []
            """if troughs[-1] < len(angle_data) - 20:
                troughs = np.insert(troughs, len(troughs), len(angle_data) - 1)"""

            return peaks, troughs
        """if troughs_num == 0:
            print("No troughs detected")
            initial_troughs = self.find_first_trough_CT_ang(angle_data)
            # troughs = np.insert(troughs, 0, initial_troughs)
            ind = self.detect_cycle_present(angle_data)
            if len(ind) == 2:
                troughs = np.insert(troughs, 0, ind[0])
                troughs = np.insert(troughs, 1, ind[1])"""
            # last_troughs = self.find_last_trough_CT_ang()
        """if peaks_num > 0 and troughs_num > 0 and peaks_num - troughs_num < 2:
            if peaks_num > troughs_num and peaks[0] < troughs[0]:
                troughs = np.insert(troughs, 0, self.find_first_trough_CT_ang(self.calculator.Calculate_derivative(angle_data)))
            if peaks_num == troughs_num and peaks[0] < troughs[0]:
                troughs = np.insert(troughs, 0, self.find_first_trough_CT_ang(self.calculator.Calculate_derivative(angle_data)))
                peaks = np.insert(peaks, len(peaks), len(angle_data) - 1)
            if peaks_num < troughs_num and peaks[0] > troughs[0]:
                peaks = np.insert(peaks, len(peaks), len(angle_data) - 1)
        elif peaks_num == 1 and troughs_num ==0:
            troughs = np.insert(troughs, 0, self.find_first_trough_CT_ang(self.calculator.Calculate_derivative(angle_data)))
            troughs = np.insert(troughs, len(troughs), len(angle_data) - 1)
        else:
            return np.array([0]), np.array([0])"""
        return peaks, troughs
    def detect_peaks_troughs(self, signal, leg):

        peak_prom = 25
        if leg[2] != "h":
            troughs, props = find_peaks(-signal, prominence=10, height=(-150, -10))
            # troughs = np.array([int(t) for t in troughs if -signal[t] >= 5])

            peaks, properties = find_peaks(signal, prominence=peak_prom)
            # peaks = np.array([int(p) for p in peaks if signal[p] >= 5])
            peaks, troughs = self.leg_search_cycle_start_classifier(peaks, troughs, signal)

            # Get the surrounding minima (left and right base of the peak)
            left_bases = properties['left_bases']
            right_bases = properties['right_bases']

            # Set threshold for minimav
            minima_threshold = 10

            # Keep only peaks whose surrounding minima are BOTH below the threshold
            selected_peaks = []
            for peak, left, right in zip(peaks, left_bases, right_bases):
                if signal[left] < minima_threshold and signal[right] < minima_threshold:
                    selected_peaks.append(peak)

            selected_peaks = np.array(selected_peaks)

        else:
            troughs, props = find_peaks(-signal, prominence=10, height=(-150, -10))

            peaks, _ = find_peaks(signal, prominence=peak_prom)

            peaks, troughs = self.leg_search_cycle_start_classifier(peaks, troughs, signal)

        return peaks, troughs
    def Detect_hard_touch(self, trial_info:Trial):
        pose_data = trial_info
        moc = trial_info.moc
        mol = trial_info.mol


        if moc < 0:
            moc = 440
            mol = moc + trial_info.fps

        center_points = self.calculator.ReadAndTranspose("platform-tip", pose_data)
        centroid, platform_direction = self.calculator.best_fit_line_3d(center_points[300:350])

        j1 = "R-mBC"
        j2 = "R-mFT"
        j3 = "R-mTT"
        j4 = "R-mLT"

        def utilities(joints, pose: Trial, platform_direction):
            angle_between_p = []
            for f in range(pose.total_frames_number):
                temp = np.asarray([pose_data.trial_data[f"{joints[1]}"].x_coord[f] -
                                   pose_data.trial_data[f"{joints[0]}"].x_coord[f],
                                   pose_data.trial_data[f"{joints[1]}"].y_coord[f] -
                                   pose_data.trial_data[f"{joints[0]}"].y_coord[f],
                                   pose_data.trial_data[f"{joints[1]}"].z_coord[f] -
                                   pose_data.trial_data[f"{joints[0]}"].z_coord[f]])
                angle_between_p.append(self.calculator.angle_between_vectors(temp, platform_direction))
            return angle_between_p
        CT_FT_angle_between_p = utilities([j1, j2], pose_data, platform_direction)
        CT_TT_angle_between_p = utilities([j1, j3], pose_data, platform_direction)
        CT_LT_angle_between_p = utilities([j1, j4], pose_data, platform_direction)


        smoothed_CT_FT = self.calculator.exponential_moving_average(CT_FT_angle_between_p, 0.01)
        smoothed_CT_TT = self.calculator.exponential_moving_average(CT_TT_angle_between_p, 0.01)
        smoothed_CT_LT = self.calculator.exponential_moving_average(CT_LT_angle_between_p, 0.01)

        normalized_CT_FT = self.calculator.normalize_list(smoothed_CT_FT)
        normalized_CT_TT = self.calculator.normalize_list(smoothed_CT_TT)
        normalized_CT_LT = self.calculator.normalize_list(smoothed_CT_LT)

        FT_baseline = np.mean(normalized_CT_FT[moc:moc + 10])
        TT_baseline = np.mean(normalized_CT_TT[moc:moc + 10])
        """top_graph_tick_value = [-0, 1]
        off_set = 0.1
        fig, ax = plt.subplots(nrows=1, ncols=1, figsize=(10, 8))
        plt.plot(normalized_CT_FT, color="blue", linewidth=3)
        plt.plot(normalized_CT_TT, color="orange", linewidth=3)
        ax.tick_params(axis="y", labelsize=15)
        ax.tick_params(axis="x", labelsize=15)
        ax.tick_params(width=3, length=5)
        ax.spines["left"].set_linewidth(2)  # Top border
        ax.spines["bottom"].set_linewidth(2)
        ax.set_xticks([0, 875, 1750], labels=["0", "3.5", "7"])
        ax.set_xlim(-100, 1850)
        ax.set_yticks([top_graph_tick_value[0], (top_graph_tick_value[0] + top_graph_tick_value[1]) / 2, top_graph_tick_value[1]])
        ax.set_ylim(top_graph_tick_value[0] - off_set, top_graph_tick_value[1] + off_set)
        sns.despine(trim=True)
        plt.ylabel("Normalized angle relative to horizontal plane",fontsize=15)
        plt.xlabel("Video duration",fontsize=15)
        plt.show()"""

        FT_drop = False
        TT_drop = False
        TT_rise = False
        for i in range(moc, len(normalized_CT_FT)):
            if ((normalized_CT_FT[i] - FT_baseline) / FT_baseline) < -0.3:
                FT_drop = True
                break
        for i in range(moc, len(normalized_CT_TT)):
            if ((normalized_CT_TT[i] - TT_baseline) / TT_baseline) < -0.3 and normalized_CT_TT[i] > 0.2:
                TT_drop = True
                TT_rise = False
                break
            if ((normalized_CT_TT[i] - TT_baseline) / TT_baseline) > 0.3 and normalized_CT_TT[i] > 0.2:
                TT_rise = True
                TT_drop = False
                break

        if FT_drop and TT_rise:
            print("Hard touch")
            return True
        elif FT_drop and TT_drop:
            print("Normal touch")
            return False
        else:
            print("Unable to tell")



        alpha = 0.5
        # return CT_FT_angle_between_p, CT_TT_angle_between_p, CT_LT_angle_between_p
        leg_stuck = self.detect_leg_stuck(CT_FT_angle_between_p, CT_TT_angle_between_p, CT_LT_angle_between_p, moc, mol, alpha)
        return leg_stuck
    def MOC_touch_type_classifier(self, joint, type, signal_to_use, trend_duration=0, analysis_duration=0):
        from scipy.stats import linregress

        if joint == "TiTa":

            if type == "slopes":
                wind = 10
                slopes = []
                for f in range(analysis_duration - wind):
                    x = np.arange(wind)
                    slope, _, _, _, _ = linregress(x, signal_to_use[f:f + wind])

                    slopes.append(slope)
                extreme_counts = len([e for e in slopes if e > 2])
                return extreme_counts > 3, extreme_counts
            if type == "trend":
                x = np.arange(trend_duration)
                general_trend, _, _, _, _ = linregress(x, signal_to_use[:trend_duration])
                return general_trend >= 0.03, general_trend
            if type == "posture":

                pass


        if joint == "CxTr":
            pass
        return None
    def all_elements_in_between(self, list1, list2):
        """
        Check if each element in list1 is between two consecutive elements in sorted list2.

        Parameters:
        ----------
        list1 : list of floats or ints
        list2 : list of floats or ints (must have len >= 2)

        Returns:
        -------
        True if all elements in list1 fall strictly between consecutive elements of list2.
        """
        list2 = sorted(list2)
        for x in list1:
            found = False
            for a, b in zip(list2, list2[1:]):
                if a < x < b:
                    found = True
                    break
            if not found:
                return False
        return True
    def detect_leg_contact(self, data):
        deri = np.gradient(data)
        for i in range(len(data)):
            if data[i] < 0.55 and deri[i] < 0.05:
                return i
        return np.nan
    def find_percent_index(self, signal, target=0.8):
        signal = np.asarray(signal)
        threshold = target * np.nanmax(signal)

        indices = np.where(signal >= threshold)[0]

        if len(indices) == 0:
            return None  # never reaches 80%

        return indices[0]

    def detect_leg_search(self, ft, ct,
                          ft_vel_thresh=1.5,
                          ct_vel_thresh=1.5,
                          min_ft_change=10,
                          pattern_duration=20,
                          idle_reset_thresh=(0.5, 0.5),
                          ct_start_thresh=100):
        """
        Detect events where FT rises followed by delayed CT rise,
        only if CT starts below a certain threshold.

        Parameters:
            ft, ct : array-like
                Angle signals for FT and CT.
            ft_vel_thresh : float
                Velocity threshold for FT rise to start pattern.
            ct_vel_thresh : float
                Velocity threshold for CT rise to complete pattern.
            min_ft_change : float
                Minimum change in FT angle to consider as real rise.
            pattern_duration : int
                Max number of frames between FT rise and CT rise.
            idle_reset_thresh : tuple(float, float)
                Velocity thresholds for FT and CT below which system resets to idle.
            ct_start_thresh : float
                Maximum CT value allowed at FT start to consider a valid rise.

        Returns:
            num_events : int
            event_indices : list of frame indices where events occurred
        """

        ft_s = np.asarray(ft)
        ct_s = np.asarray(ct)

        dft = np.gradient(ft_s)
        dct = np.gradient(ct_s)

        state = "idle"
        event_indices = []
        ft_start = None

        for i in range(len(dft)):

            if state == "idle":
                # FT begins rising strongly, CT not rising yet, and CT below threshold
                if dft[i] > ft_vel_thresh and dct[i] <= ct_vel_thresh and ct_s[i] <= ct_start_thresh:
                    state = "ft_rising"
                    ft_start = i

            elif state == "ft_rising":
                # ensure FT actually moved enough
                if ft[i] - ft[ft_start] < min_ft_change:
                    continue

                # Check duration threshold
                if i - ft_start > pattern_duration:
                    state = "idle"
                    ft_start = None
                    continue

                # CT rises after FT
                if dct[i] > ct_vel_thresh:
                    event_indices.append(i)
                    state = "event"

            elif state == "event":
                # wait until movement settles before detecting another
                if dft[i] < idle_reset_thresh[0] and dct[i] < idle_reset_thresh[1]:
                    state = "idle"
                    ft_start = None

        return len(event_indices), event_indices
    def detect_landing(self, data, windsize=5):
        data = self.calculator.normalize_list(data)
        from scipy.ndimage import gaussian_filter1d

        data = gaussian_filter1d(data, sigma=10)
        for i in range(len(data) - windsize):
            if np.mean(data[i:i + windsize]) < 0.2:
                return i
        return -1

"""
These functions are responsible for manipulating files
"""
class FileManipulation:
    # Extract the data path of the csv files and organize them in the order by fly number
    def Write_to_csv(self, LandingProbData, LandingLatencyData, output_path, compare_path):
        Manual_data = pd.read_excel(compare_path)
        T = ["Trial_" + str(i) for i in range(1, 21)]
        Manual_data = Manual_data[T]
        # Create an Excel writer using openpyxl
        excel_writer = pd.ExcelWriter(output_path, engine='openpyxl')
        LandingProbData.to_excel(excel_writer, index=False, sheet_name='Sheet1')
        # Access the workbook and sheet
        workbook = excel_writer.book
        sheet = workbook['Sheet1']

        # Define cell colors
        yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
        orange_fill = PatternFill(start_color='FFC000', end_color='FFC000', fill_type='solid')
        red_fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
        blue_fill = PatternFill(start_color='9BC2E6', end_color='9BC2E6', fill_type='solid')
        gray_fill = PatternFill(start_color='696969', end_color='696969', fill_type='solid')

        # Apply cell colors based on conditions
        for row_idx in range(2, sheet.max_row + 1):  # Start from row 2 (excluding header)
            for col_idx in range(2, sheet.max_column + 1):  # Start from column 2 (excluding Fly# column)
                cell_value = sheet.cell(row=row_idx, column=col_idx).value
                cell = sheet.cell(row=row_idx, column=col_idx)
                if cell_value == 1:
                    if isinstance(Manual_data[T].iloc[row_idx - 2][col_idx - 2], str):
                        cell.fill = gray_fill
                    else:
                        if Manual_data[T].iloc[row_idx - 2][col_idx - 2] > 1:
                            cell.fill = yellow_fill
                        else:
                            cell.fill = gray_fill
                    cell.value = LandingLatencyData[row_idx - 2][col_idx - 2]
                elif cell_value == -1:
                    if Manual_data[T].iloc[row_idx - 2][col_idx - 2] == -1:
                        cell.fill = orange_fill
                    else:
                        cell.fill = gray_fill
                elif cell_value == 0:
                    if isinstance(Manual_data[T].iloc[row_idx - 2][col_idx - 2], str):
                        if Manual_data[T].iloc[row_idx - 2][col_idx - 2] == "NotFlying":
                            cell.fill = blue_fill
                        else:
                            cell.fill = gray_fill
                    else:
                        cell.fill = gray_fill
                    cell.value = "NotFlying"
                elif isinstance(type(cell_value), type(str)):
                    # print("N/A")
                    # print(cell_value)
                    if pd.isna(Manual_data[T].iloc[row_idx - 2][col_idx - 2]):
                        cell.fill = red_fill
                    else:
                        cell.fill = gray_fill
                    cell.value = 'N/A'
        # Save the Excel file
        excel_writer.save()
        print("Excel file has been created with modifications.")
    def OutptuHardtouchPrediction(self, Predicted_Data, outputpath):
        df_transformed = pd.DataFrame(Predicted_Data, columns=[f"Trial_{i + 1}" for i in range(20)])
        df_transformed.insert(0, "Fly#", range(1, len(df_transformed) + 1))

        # Save to an Excel file
        output_file = f"{outputpath}.xlsx"
        df_transformed.to_excel(output_file, index=False)

        # Load the workbook for formatting
        wb = load_workbook(output_file)
        ws = wb.active

        # Define color fills
        color_fills = {
            "-1": PatternFill(start_color="FF9933", end_color="FF9933", fill_type="solid"),
            "number": PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
            # Yellow for other numbers
        }

        # Apply color coding (skip header row)
        for row in ws.iter_rows(min_row=2, min_col=2):  # Start from column 2 to skip "Fly#"
            for cell in row:
                if isinstance(cell.value, (int, float)) and cell.value == -1:
                    cell.fill = color_fills["-1"]
                elif isinstance(cell.value, (int, float)):  # Any other number
                    cell.fill = color_fills["number"]

        # Save the formatted workbook
        wb.save(output_file)

        print(f"File saved as {output_file}")
    def OutptuPrediction(self, Predicted_Data, outputpath):
        df_transformed = pd.DataFrame(Predicted_Data, columns=[f"Trial_{i + 1}" for i in range(20)])
        df_transformed.insert(0, "Fly#", range(1, len(df_transformed) + 1))

        # Replace NaN values with "N/A" for display
        df_transformed.fillna("N/A", inplace=True)

        # Save to an Excel file
        output_file = f"{outputpath}.xlsx"
        df_transformed.to_excel(output_file, index=False)

        # Load the workbook for formatting
        wb = load_workbook(output_file)
        ws = wb.active

        # Define color fills
        color_fills = {
            "N/A": PatternFill(start_color="FF6666", end_color="FF6666", fill_type="solid"),  # Red for N/A
            "NF": PatternFill(start_color="9BC2E6", end_color="9BC2E6", fill_type="solid"),  # Blue for NF
            "-1": PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid"),  # Orange for -1
            "OT": PatternFill(start_color="92D050", end_color="92D050", fill_type="solid"),
            "number": PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
            # Yellow for other numbers
        }

        # Apply color coding (skip header row)
        for row in ws.iter_rows(min_row=2, min_col=2):  # Start from column 2 to skip "Fly#"
            for cell in row:
                if isinstance(cell.value, str) and cell.value == "N/A":
                    cell.fill = color_fills["N/A"]
                elif isinstance(cell.value, str) and cell.value == "NF":
                    cell.fill = color_fills["NF"]
                elif isinstance(cell.value, (int, float)) and cell.value == -1:
                    cell.fill = color_fills["-1"]
                elif isinstance(cell.value, str) and cell.value == "OT":
                    cell.fill = color_fills["OT"]
                elif isinstance(cell.value, (int, float)):  # Any other number
                    cell.fill = color_fills["number"]

        # Save the formatted workbook
        wb.save(output_file)

        print(f"File saved as {output_file}")
    def addColumns_to_file(self, data, filename, group):
        if os.path.exists(f"{filename}.csv"):
            df = pd.read_csv(f"{filename}.csv")

            if len(data) < len(df):
                data += [np.nan] * (len(df) - len(data))
            elif len(data) > len(df):
                extra_rows = len(data) - len(df)
                df = df.reindex(range(len(df) + extra_rows))

            df[group] = data
        else:
            df = pd.DataFrame({group: data})
        df.to_csv(f"{filename}.csv", index=False)

    def read_secondary_contact_data(self, data, legs, filepath=None):
        if filepath is not None:
            data_to_read = pd.read_csv(filepath)
        else:
            data_to_read = data
        successful_landing_data = dict()
        failed_landing_data = dict()
        jointA = "FT"
        jointB = "TT"
        for row in data_to_read.iterrows():
            for l in legs:
                if l not in successful_landing_data.keys() and l not in failed_landing_data.keys():
                    successful_landing_data[l] = []
                    failed_landing_data[l] = []

                if row[1][l + jointA] > 0 and int(row[1][l + jointA]) != 10000:

                    if row[1]["Result"] != "Failed":
                        successful_landing_data[l].append(row[1][l + jointA])
                    else:
                        failed_landing_data[l].append(row[1][l + jointA])
                elif row[1][l + jointB] > 0 and int(row[1][l + jointB]) != 10000:

                    if row[1]["Result"] != "Failed":
                        successful_landing_data[l].append(row[1][l + jointB])
                    else:
                        failed_landing_data[l].append(row[1][l + jointB])
                else:

                    if row[1]["Result"] != "Failed":
                        successful_landing_data[l].append(np.nan)
                    else:
                        failed_landing_data[l].append(np.nan)
        successful_landing_data = pd.DataFrame(successful_landing_data)
        failed_landing_data = pd.DataFrame(failed_landing_data)
        return successful_landing_data, failed_landing_data
    def read_leg_search_data(self, data, legs, filepath=None):
        if filepath is not None:
            data_to_read = pd.read_csv(filepath)
        else:
            data_to_read = data
        failed_landing = data_to_read[data_to_read["Result"] == "Failed"]
        failed_landing = failed_landing[legs]

        successful_landing = data_to_read[data_to_read["Result"] == "Success"]
        successful_landing = successful_landing[legs]

        return successful_landing, failed_landing
"""
These functions are responsible for analyzing the fly group data
"""
class GroupDataAnalyzer:
    def __init__(self, platform_offset, radius, FPS):
        self.fps = FPS
        self.detector = DetectCharacteristics(radius, FPS)
        self.calculator = SimpleCalculation()
        self.manipulator = FileManipulation()
        self.offset = platform_offset
    def DetermineTiTaMOC(self, trial_info:Trial):
        start = int(trial_info.fps)
        end = int(trial_info.fps * 4)

        center_points = self.calculator.ReadAndTranspose("platform-tip", trial_info)
        R_mTT_points = self.calculator.ReadAndTranspose("R-mTT", trial_info)
        R_mLT_points = self.calculator.ReadAndTranspose("R-mLT", trial_info)
        L_mTT_points = self.calculator.ReadAndTranspose("L-mTT", trial_info)
        L_mLT_points = self.calculator.ReadAndTranspose("L-mLT", trial_info)
        contact_count_L = 0
        contact_count_R = 0
        Contact_threshold = 1

        centroid, direction = self.calculator.best_fit_line_3d(center_points[300:350])
        for frame in range(start, end):
            # print(frame)
            Intersect_R, position_R = self.detector.check_leg_platform_intersection(R_mTT_points[frame],
                                                                                    R_mLT_points[frame],
                                                                                    direction,
                                                                                    center_points[frame],
                                                                                    self.offset)

            Intersect_L, position_L = self.detector.check_leg_platform_intersection(L_mTT_points[frame],
                                                                                    L_mLT_points[frame],
                                                                                    direction,
                                                                                    center_points[frame],
                                                                                    self.offset)

            if Intersect_L:
                contact_count_L += 1
            else:
                contact_count_L = 0
            if contact_count_L >= Contact_threshold:
                return frame, "L", position_L

            if Intersect_R:
                contact_count_R += 1
            else:
                contact_count_R = 0
            if contact_count_R >= Contact_threshold:
                return frame, "R", position_R
        return 0, "NoContact", None

    def Determine_all_flying_posture(self, group_info:Group, filter_high_latency=False):
        angles = [["R-mCT", "R-mFT", "R-mTT"], ["L-mCT", "L-mFT", "L-mTT"]]
        for index in group_info.get_targeted_trials(["Landing", "Flying"]):
            joints_angle = self.calculator.Calculate_joint_angle(group_info.fly_kinematic_data[f"F{index[0]}T{index[1]}"], angles)
            if filter_high_latency:
                if self.detector.detect_stable_posture(joints_angle):
                    group_info.fly_kinematic_data[f"F{index[0]}T{index[1]}"].L_stable_FT_angle = np.mean(joints_angle["L-mFT"][:group_info.fps[0]])
                    group_info.fly_kinematic_data[f"F{index[0]}T{index[1]}"].R_stable_FT_angle = np.mean(joints_angle["R-mFT"][:group_info.fps[0]])
                else:
                    group_info.fly_kinematic_data[f"F{index[0]}T{index[1]}"].L_stable_FT_angle = None
                    group_info.fly_kinematic_data[f"F{index[0]}T{index[1]}"].R_stable_FT_angle = None
            else:
                group_info.fly_kinematic_data[f"F{index[0]}T{index[1]}"].L_stable_FT_angle = np.mean(joints_angle["L-mFT"][:group_info.fps[0]])
                group_info.fly_kinematic_data[f"F{index[0]}T{index[1]}"].R_stable_FT_angle = np.mean(joints_angle["R-mFT"][:group_info.fps[0]])

        # self.manipulator.OutptuPrediction(collected_stable_posture, f"{groupname}_StablePosture_CT")
    def TiTa_relative_contact(self, group_info:Group):
        Contact_relative_position = []
        Latency = []
        for index in group_info.get_targeted_trials("Landing"):
            print(f"Determine TiTa contact point for fly {index[0]}", f"Trial {index[1]}")
            pose_data = group_info.fly_kinematic_data[f"F{index[0]}T{index[1]}"]
            moc, leg, tarsus_contact_position =  self.DetermineTiTaMOC(pose_data)
            if leg == "R":
                if (group_info.ll_data.iloc[index[0] - 1][index[1] - 1] / group_info.fps) < 1:
                    Contact_relative_position.append(tarsus_contact_position)
                    Latency.append(group_info.ll_data.iloc[index[0] - 1][index[1] - 1] / group_info.fps)
        return Contact_relative_position, Latency

    def Calculate_angle_traces(self, group_info: Group, index_to_iterate, angles, threshold=None,
                               start=-0.3, end=0.5, normalize_time=False, normalize_angle=False):


        previous_fly = index_to_iterate[0][0]
        collected_data = []
        indi_fly_data = {}

        for index in index_to_iterate:
            trial_key = f"F{index[0]}T{index[1]}"
            trial_info = group_info.fly_kinematic_data[trial_key]

            if index[0] != previous_fly:
                if indi_fly_data:
                    collected_data.append(indi_fly_data)
                indi_fly_data = {}
                previous_fly = index[0]

            MOC = trial_info.moc
            MOL = trial_info.mol

            if MOC <= 0:
                continue

            angs = self.calculator.Calculate_joint_angle(trial_info, angles)

            for joint in angles:
                joint_name = joint[1]

                if joint_name not in indi_fly_data:
                    indi_fly_data[joint_name] = []

                if normalize_time:
                    Joint_signal = np.asarray(angs[joint_name][int(MOC):int(MOL)])
                    Joint_signal = self.calculator.Normalized_time(Joint_signal)
                else:
                    Joint_signal = np.asarray(angs[joint_name][int(MOC) - int(-start * trial_info.fps): int(MOC) + int(end * trial_info.fps)])

                    if trial_info.fps == 200:
                        target_len = int(round((end - start) * 250))
                        Joint_signal = self.calculator.Normalized_time(Joint_signal, target_len)

                if normalize_angle:
                    Joint_signal = self.calculator.normalize_list(Joint_signal)

                if threshold is None:
                    if MOL == -1:
                        indi_fly_data[joint_name].append((-1, index, 250, Joint_signal))
                    elif (MOL - MOC) / trial_info.fps <= 0.3:
                        indi_fly_data[joint_name].append((0.3, index, 250, Joint_signal))
                    elif (MOL - MOC) / trial_info.fps <= 0.683:
                        indi_fly_data[joint_name].append((0.68, index, 250, Joint_signal))
                    else:
                        indi_fly_data[joint_name].append((1, index, 250, Joint_signal))
                else:
                    if MOL > 0 and ((MOL - MOC) / trial_info.fps) <= threshold:
                        indi_fly_data[joint_name].append((threshold, index, 250, Joint_signal))
                    else:
                        indi_fly_data[joint_name].append((-1, index, 250, Joint_signal))

        if indi_fly_data:
            collected_data.append(indi_fly_data)

        return collected_data

    def AnalyzeSecondaryContact(self, index_to_iterate, group_info:Group, threshold, filename=""):
        global ax_flip_count
        global total_count
        segs = [["L-fFT", "L-fTT", 0.45],
                ["L-fTT", "L-fLT", 0.4],

                ["L-mFT", "L-mTT", 0.5],
                ["L-mTT", "L-mLT", 0.4],

                ["L-hFT", "L-hTT", 0.6],
                ["L-hTT", "L-hLT", 0.5],

                ["R-fFT", "R-fTT", 0.45],
                ["R-fTT", "R-fLT", 0.4],

                ["R-mFT", "R-mTT", 0.5],
                ["R-mTT", "R-mLT", 0.4],

                ["R-hFT", "R-hTT", 0.5],
                ["R-hTT", "R-hLT", 0.4]]

        pts = ["L-fFT", "L-fTT", "L-fLT",
               "L-mFT", "L-mTT", "L-mLT",
               "L-hFT", "L-hTT", "L-hLT",
               "R-fFT", "R-fTT", "R-fLT",
               "R-mFT", "R-mTT", "R-mLT",
               "R-hFT", "R-hTT", "R-hLT"]

        indi_leg_contact_event = dict()
        for j in segs:
            indi_leg_contact_event[j[0]] = []

        indi_leg_contact_event["Index"] = []
        indi_leg_contact_event["Result"] = []

        for i, index in enumerate(index_to_iterate):
            pose_data = group_info.fly_kinematic_data[f"F{index[0]}T{index[1]}"]

            start = int(pose_data.moc)
            end = int(pose_data.mol)
            if start > 0:
                if end == -1:
                    end = start + int(threshold * pose_data.fps)
                    indi_leg_contact_event["Result"].append("Failed")
                elif (end - start) / pose_data.fps > threshold:
                    end = start + int(threshold * pose_data.fps)
                    indi_leg_contact_event["Result"].append("Failed")
                elif (end - start) / pose_data.fps <= threshold:
                    indi_leg_contact_event["Result"].append("Success")
                else:
                    print("Unable to categorize!")
            line_points, plane_points, verts, cylinder_top, cylinder_bottom, direction, perp_vector1, perp_vector2 = (
                self.calculator.transform_coords_and_calculate_platform_data(trial_info=pose_data,
                                                                             platform_offset=0.03,
                                                                             platform_height=3,
                                                                             radius=0.35))

            Original_points = dict()
            center_points = self.calculator.ReadAndTranspose("platform-tip", pose_data)[start:end]
            for p in pts:
                Original_points[p] = self.calculator.ReadAndTranspose(p, pose_data)[start:end]
            # event_recorder = []
            for point in segs:
                NoContact = True
                stable_contact = 2
                for current_frame in range(end - start):
                    A = Original_points[point[0]][current_frame]
                    B = Original_points[point[1]][current_frame]
                    P1 = center_points[current_frame]

                    d = direction
                    r = point[2]
                    h = 3

                    intersects_side, pt_side, min_dist = self.calculator.check_cylinder_side_intersection(A, B, P1, d, r, h)
                    intersects_top, pt_top = self.detector.check_leg_platform_intersection(A, B, d, center_points[current_frame], 0.03)
                    if intersects_top or intersects_side:
                        stable_contact += 1
                    else:
                        stable_contact = 0
                    if stable_contact >= 1:
                        indi_leg_contact_event[point[0]].append(current_frame / pose_data.fps)
                        NoContact = False
                        break

                if NoContact:
                    indi_leg_contact_event[point[0]].append(10000)

            if index not in indi_leg_contact_event["Index"]:
                indi_leg_contact_event["Index"].append(index)
        indi_leg_contact_event = pd.DataFrame(indi_leg_contact_event)
        indi_leg_contact_event.to_csv(f"{group_info.group_name}-{filename}-{threshold}.csv")
        return indi_leg_contact_event

    def Calculate_contact_leg_metrices(self, group_info:Group, index_to_iterate, joint_angle, threshold=0.71):
        joint_to_examine = joint_angle[0][1]

        failed_ang_v = []
        failed_ang_proj = []
        failed_ang_ft = []

        success_ang_v = []
        success_ang_proj = []
        success_ang_ft = []

        for index in index_to_iterate:
            trial_info = group_info.fly_kinematic_data[f"F{index[0]}T{index[1]}"]
            start = trial_info.moc
            if start > 0:
                line_points, plane_points, verts, cylinder_top, cylinder_bottom, direction, perp_vector1, perp_vector2 = (
                    self.calculator.transform_coords_and_calculate_platform_data(trial_info=trial_info,
                                                                                 platform_offset=0.03,
                                                                                 platform_height=3,
                                                                                 radius=0.45))

                angs = self.calculator.Calculate_joint_angle(trial_info, joint_angle)
                Failed = False
                analysis_start = trial_info.moc
                analysis_end = None

                if trial_info.mol > 0 and ((trial_info.mol - trial_info.moc) / trial_info.fps) <= threshold:  # Successful landing
                    analysis_end = trial_info.mol
                elif trial_info.mol < 0 or (trial_info.mol > 0 and ((trial_info.mol - trial_info.moc) / trial_info.fps) > threshold):  # Failed landing
                    analysis_end = trial_info.moc + int(threshold * trial_info.fps)
                    Failed = True
                else:
                    analysis_end = trial_info.moc + int(threshold * trial_info.fps)
                    print("Uncategorized!")

                L_mBC = self.calculator.get_stable_point(trial_info, "L-mBC", analysis_start, 10)
                R_mBC = self.calculator.get_stable_point(trial_info, "R-mBC", analysis_start, 10)
                R_mTT = self.calculator.get_stable_point(trial_info, "R-mTT", analysis_start, 10)
                R_mLT = self.calculator.get_stable_point(trial_info, "R-mLT", analysis_start, 10)

                ang_v = np.mean(np.gradient(angs[joint_to_examine][analysis_start:analysis_end])) * 250
                proj_ang = 180 - self.calculator.projected_signed_angle(R_mTT, R_mLT, L_mBC, R_mBC, direction)
                ft_ang = np.mean(angs[joint_to_examine][analysis_start - int(0.5 * trial_info.fps):analysis_start])

                if Failed:
                    failed_ang_v.append(ang_v)
                    failed_ang_proj.append(proj_ang)
                    failed_ang_ft.append(ft_ang)
                else:
                    success_ang_v.append(ang_v)
                    success_ang_proj.append(proj_ang)
                    success_ang_ft.append(ft_ang)

        return success_ang_v, success_ang_ft, success_ang_proj, failed_ang_v, failed_ang_ft, failed_ang_proj

    def Analyze_leg_search(self, group_info:Group, index_to_iterate=None, filename="", threshold=0.71):
        Angles = [[["L-fBC", "L-fCT", "L-fFT"], ["L-fCT", "L-fFT", "L-fTT"]],
                  [["L-mBC", "L-mCT", "L-mFT"], ["L-mCT", "L-mFT", "L-mTT"]],
                  [["L-hBC", "L-hCT", "L-hFT"], ["L-hCT", "L-hFT", "L-hTT"]],
                  [["R-fBC", "R-fCT", "R-fFT"], ["R-fCT", "R-fFT", "R-fTT"]],
                  [["R-mBC", "R-mCT", "R-mFT"], ["R-mCT", "R-mFT", "R-mTT"]],
                  [["R-hBC", "R-hCT", "R-hFT"], ["R-hCT", "R-hFT", "R-hTT"]]]

        # self.angles = [["L-fBC", "L-fCT", "L-fFT"], ["L-fCT", "L-fFT", "L-fTT"]]
        leg_search_data = dict()
        leg_search_data["L-f"] = []
        leg_search_data["L-m"] = []
        leg_search_data["L-h"] = []
        leg_search_data["R-f"] = []
        leg_search_data["R-m"] = []
        leg_search_data["R-h"] = []
        leg_search_data["Index"] = []
        leg_search_data["Result"] = []

        if index_to_iterate is None:
            index_to_iterate = group_info.get_targeted_trials(["Landing"])

        for index in index_to_iterate:
            trial_info = group_info.fly_kinematic_data[f"F{index[0]}T{index[1]}"]
            start = trial_info.moc
            end = trial_info.mol
            leg_search_data["Index"].append(index)
            if end == -1:
                end = start + int(threshold * trial_info.fps)
                leg_search_data["Result"].append("Failed")
            elif (end - start) / trial_info.fps > threshold:
                end = start + int(threshold * trial_info.fps)
                leg_search_data["Result"].append("Failed")
            elif (end - start) / trial_info.fps <= threshold:
                leg_search_data["Result"].append("Success")
            else:
                print("Unable to categorize!")

            for pair in Angles:
                ags = self.calculator.Calculate_joint_angle(trial_info, pair)
                ct_trace = None
                ft_trace = None
                for ag in pair:
                    trace = np.array(ags[ag[1]][start:end + 1])
                    if "CT" in ag[1]:
                        ct_trace = trace
                    if "FT" in ag[1]:
                        ft_trace = trace
                if "h" not in pair[0][0][:3]: # not hind leg
                    counts, events = self.detector.detect_leg_search(ft_trace, ct_trace)
                    leg_search_data[pair[0][0][:3]].append(counts)
                else:
                    peaks, _ = find_peaks(ct_trace, prominence=15)
                    leg_search_data[pair[0][0][:3]].append(len(peaks))
        leg_search_data = pd.DataFrame(leg_search_data)
        leg_search_data.to_csv(f"{group_info.group_name}-{filename}-LS_data_.csv")
        return leg_search_data
    def Analyze_leg_search_CHR(self, group_info:Group, index_to_iterate=None, filename="", threshold=0.71):
        Angles = [[["L-fBC", "L-fCT", "L-fFT"], ["L-fCT", "L-fFT", "L-fTT"]],
                  [["L-mBC", "L-mCT", "L-mFT"], ["L-mCT", "L-mFT", "L-mTT"]],
                  [["L-hBC", "L-hCT", "L-hFT"], ["L-hCT", "L-hFT", "L-hTT"]],
                  [["R-fBC", "R-fCT", "R-fFT"], ["R-fCT", "R-fFT", "R-fTT"]],
                  [["R-mBC", "R-mCT", "R-mFT"], ["R-mCT", "R-mFT", "R-mTT"]],
                  [["R-hBC", "R-hCT", "R-hFT"], ["R-hCT", "R-hFT", "R-hTT"]]]

        # self.angles = [["L-fBC", "L-fCT", "L-fFT"], ["L-fCT", "L-fFT", "L-fTT"]]
        leg_search_data = dict()
        leg_search_data["L-f"] = []
        leg_search_data["L-m"] = []
        leg_search_data["L-h"] = []
        leg_search_data["R-f"] = []
        leg_search_data["R-m"] = []
        leg_search_data["R-h"] = []
        leg_search_data["Index"] = []
        leg_search_data["Result"] = []

        if index_to_iterate is None:
            index_to_iterate = group_info.get_targeted_trials(["Landing"])

        for index in index_to_iterate:
            trial_info = group_info.fly_kinematic_data[f"F{index[0]}T{index[1]}"]
            angs = self.calculator.Calculate_joint_angle(trial_info, [["L-wing", "L-wing-hinge", "R-wing"]])


            start = 750
            end = self.detector.detect_landing(angs["L-wing-hinge"][750:1250]) + start
            plt.plot(angs["L-wing-hinge"][750:1250])
            if end == 749:
                end = 1250

            leg_search_data["Index"].append(index)
            if end == 1250:
                end = start + int(threshold * trial_info.fps)
                leg_search_data["Result"].append("Failed")
            elif (end - start) / trial_info.fps > threshold:
                end = start + int(threshold * trial_info.fps)
                leg_search_data["Result"].append("Failed")
            elif (end - start) / trial_info.fps <= threshold:
                leg_search_data["Result"].append("Success")
            else:
                print("Unable to categorize!")
            print(index, end)

            for pair in Angles:
                ags = self.calculator.Calculate_joint_angle(trial_info, pair)
                ct_trace = None
                ft_trace = None
                for ag in pair:
                    trace = np.array(ags[ag[1]][start:end + 1])
                    if "CT" in ag[1]:
                        ct_trace = trace
                    if "FT" in ag[1]:
                        ft_trace = trace
                if "h" not in pair[0][0][:3]: # not hind leg
                    counts, events = self.detector.detect_leg_search(ft_trace, ct_trace)
                    leg_search_data[pair[0][0][:3]].append(counts)
                else:
                    peaks, _ = find_peaks(ct_trace, prominence=15)
                    leg_search_data[pair[0][0][:3]].append(len(peaks))
        leg_search_data = pd.DataFrame(leg_search_data)
        leg_search_data.to_csv(f"{group_info.group_name}-{filename}-LS_data_.csv")
        return leg_search_data
    def combine_data(self, group_info:Group, type, Opto, group1_index=None, group2_index=None):
        indi_legs = ["L-f", "L-m", "L-h"]
        if group1_index is not None:
            combined_IT = None
            combined_OT = None
            combined_Success = None
            combined_Failed = None
            IT_Success = None
            IT_Failed = None
            OT_Success = None
            OT_Failed = None
            if type == "LS":
                Success_IT, Failed_IT = self.manipulator.read_leg_search_data(self.Analyze_leg_search(group_info, group1_index, group_info.group_name), indi_legs)
                Success_OT, Failed_OT = self.manipulator.read_leg_search_data(self.Analyze_leg_search(group_info, group2_index, group_info.group_name), indi_legs)

                combined_IT = pd.concat([Success_IT, Failed_IT])
                combined_IT = combined_IT.sum(axis=1).to_frame(name="sum")
                combined_OT = pd.concat([Success_OT, Failed_OT])
                combined_OT = combined_OT.sum(axis=1).to_frame(name="sum")

                combined_Success = pd.concat([Success_IT, Success_OT])
                combined_Success = combined_Success.sum(axis=1).to_frame(name="sum")
                combined_Failed = pd.concat([Failed_IT, Failed_OT])
                combined_Failed = combined_Failed.sum(axis=1).to_frame(name="sum")

                IT_Success = Success_IT.sum(axis=1).to_frame(name="sum")
                IT_Failed = Failed_IT.sum(axis=1).to_frame(name="sum")
                OT_Success = Success_OT.sum(axis=1).to_frame(name="sum")
                OT_Failed = Failed_OT.sum(axis=1).to_frame(name="sum")

            if type == "SC":
                Success_IT, Failed_IT = self.manipulator.read_secondary_contact_data(self.AnalyzeSecondaryContact(group1_index, group_info,  0.71), indi_legs)
                Success_OT, Failed_OT = self.manipulator.read_secondary_contact_data(self.AnalyzeSecondaryContact(group2_index, group_info, 0.71), indi_legs)

                combined_IT = pd.concat([Success_IT, Failed_IT])
                combined_IT = combined_IT.min(axis=1).dropna().to_frame(name="SC")
                combined_OT = pd.concat([Success_OT, Failed_OT])
                combined_OT = combined_OT.min(axis=1).dropna().to_frame(name="SC")

                combined_Success = pd.concat([Success_IT, Success_OT])
                combined_Success = combined_Success.min(axis=1).dropna().to_frame(name="SC")
                combined_Failed = pd.concat([Failed_IT, Failed_OT])
                combined_Failed = combined_Failed.min(axis=1).dropna().to_frame(name="SC")

                IT_Success = Success_IT.min(axis=1).dropna().to_frame(name="SC")
                IT_Failed = Failed_IT.min(axis=1).dropna().to_frame(name="SC")
                OT_Success = Success_OT.min(axis=1).dropna().to_frame(name="SC")
                OT_Failed = Failed_OT.min(axis=1).dropna().to_frame(name="SC")

            return combined_IT, combined_OT, combined_Success, combined_Failed, IT_Success, IT_Failed, OT_Success, OT_Failed

        if not Opto:
            index_to_iterate = group_info.get_targeted_trials(["Landing", "Flying"])
            combined = None
            threshold = 0.71
            if type == "LS":

                Success, Failed = self.manipulator.read_leg_search_data(self.Analyze_leg_search(group_info, index_to_iterate, group_info.group_name, threshold), indi_legs)
                combined = pd.concat([Success, Failed])
                combined = combined.sum(axis=1).to_frame(name="sum")
            if type == "SC":
                Success, Failed = self.manipulator.read_secondary_contact_data(self.AnalyzeSecondaryContact(index_to_iterate, group_info,  threshold), indi_legs)
                combined = pd.concat([Success, Failed])
                combined = combined.min(axis=1).dropna().to_frame(name="SC")
            return combined
        else:
            LO_index = []
            NL_index = []
            index_to_iterate = group_info.get_targeted_trials(["Landing", "Flying"])
            for index in index_to_iterate:
                path = group_info.fly_kinematic_data_path[f"F{index[0]}T{index[1]}"]
                if "_NL_" in path or "OFF" in path:
                    NL_index.append(index)
                if "_LO_" in path or "ON" in path:
                    LO_index.append(index)
            ON_combined = None
            OFF_combined = None
            if type == "LS":
                Success, Failed = self.manipulator.read_leg_search_data(self.Analyze_leg_search(group_info, LO_index, "ON"), indi_legs)
                ON_combined = pd.concat([Success, Failed])
                ON_combined = ON_combined.sum(axis=1).to_frame(name="sum")

                Success, Failed = self.manipulator.read_leg_search_data(self.Analyze_leg_search(group_info, NL_index, "OFF"), indi_legs)
                OFF_combined = pd.concat([Success, Failed])
                OFF_combined = OFF_combined.sum(axis=1).to_frame(name="sum")

            if type == "SC":

                Success, Failed = self.manipulator.read_secondary_contact_data(self.AnalyzeSecondaryContact(LO_index, group_info,  0.71, "ON"), indi_legs)
                ON_combined = pd.concat([Success, Failed])
                ON_combined = ON_combined.min(axis=1).dropna().to_frame(name="SC")

                Success, Failed = self.manipulator.read_secondary_contact_data(self.AnalyzeSecondaryContact(NL_index, group_info, 0.71, "OFF"), indi_legs)
                OFF_combined = pd.concat([Success, Failed])
                OFF_combined = OFF_combined.min(axis=1).dropna().to_frame(name="SC")
            return ON_combined, OFF_combined

